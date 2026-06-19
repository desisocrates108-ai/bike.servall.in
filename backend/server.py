from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import logging
import uuid
import bcrypt
import jwt
import requests
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Any, Dict

from fastapi import (
    FastAPI, APIRouter, Depends, HTTPException, Request, Response,
    UploadFile, File, Form, Query, Header
)
from fastapi.responses import Response as FastResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr

# ============================================================
# Config
# ============================================================

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALG = "HS256"
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")
APP_NAME = "twowheeler-crm"
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Two-Wheeler Dealership CRM")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("crm")

# ============================================================
# Constants
# ============================================================

LEAD_SOURCES = ["Walk-in", "Tele-in", "Digital Marketing", "Social Media",
                "WhatsApp", "Email", "Referral", "Offline", "Cold Calling"]
PRIORITIES = ["Hot", "Warm", "Cold"]
STAGES = ["Inquiry", "Follow-up", "Hold", "Booking", "Delivery",
          "Allotment", "Feedback", "Lost"]
# Backward-compat mapping for legacy stage names (used during migration + read shim)
STAGE_ALIAS = {
    "Interest": "Follow-up",
    "Deal": "Booking",
    "Test Ride": "Follow-up",
    "Booking Hold": "Hold",
    "RTO": "Allotment",
    "Delivered": "Delivery",
    "Registration": "Allotment",
}

CUSTOMER_TYPES = ["Instant Buyer", "Token Finance Buyer", "Just Inquiry"]
PAYMENT_MODES = ["Cash", "UPI", "Bank Transfer", "Cheque", "Finance"]
FOLLOWUP_TYPES = ["Call", "WhatsApp", "Visit", "Test Ride", "Other"]
CALL_STATUSES = ["Connected", "Not Connected"]
CUSTOMER_RESPONSES = ["Interested", "Not Interested", "Call Later",
                      "Not Reachable", "Switched Off"]
OUTCOME_TAGS = ["Progressed", "No Progress", "Converted", "Lost"]
LOST_REASONS = ["Price Issue", "Competitor", "Stock Issue", "No Follow-up",
                "Not Interested", "Other"]
DEAL_LOSS_REASONS = ["Price too high", "Better competitor offer",
                     "Finance rejected", "Delay / follow-up issue",
                     "Stock Issue", "Not interested", "Other"]
DEAL_STATUSES = ["Draft", "Negotiation", "Approval Pending",
                 "Approved", "Rejected", "Converted"]
BOOKING_STATUSES = ["Pending", "Confirmed", "Cancelled"]
ALLOTMENT_STATUSES = ["Pending", "Allotted"]
LOAN_STATUSES = ["Pending", "Approved", "Rejected"]
DOC_TYPES = ["Aadhaar Card", "PAN Card", "RC Book", "Bank Passbook",
             "Finance Document", "Insurance Copy", "Invoice",
             "RTO Form 20", "RTO Form 21", "RTO Form 22",
             "Sale Challan", "Address Proof", "Photo", "Other"]
DOC_STATUSES = ["Pending", "Verified", "Rejected"]
DELIVERY_STATUSES = ["Scheduled", "Ready", "Delivered", "Cancelled"]
DEFAULT_ACCESSORIES = ["Helmet", "Ceramic Coating", "Mud Flaps",
                       "Side Guard", "Mobile Holder", "Other"]
PAYMENT_TYPES = ["Booking", "Margin", "Final", "Other"]
PAYMENT_STATUSES = ["Pending", "Partial", "Completed"]
FINANCE_STATUSES = ["Not Applied", "Applied", "Under Review", "Approved", "Rejected"]
EXCHANGE_CONDITIONS = ["Good", "Average", "Poor"]
MARGIN_ALERT_DAYS = 3

# Module 11 — WhatsApp Automation
WA_EVENTS = [
    "inquiry_created", "stage_changed", "followup_due", "deal_updated",
    "booking_confirmed", "delivery_scheduled", "delivery_completed",
    "feedback_reminder", "rc_reminder", "lost_reengage", "manual", "campaign",
]
WA_MESSAGE_TYPES = ["text", "image", "pdf", "document"]
WA_STATUSES = ["PENDING", "SENT", "DELIVERED", "READ", "FAILED"]
WA_REPLY_TAGS = ["Interested", "Not Interested", "Call Back", "Other"]
WA_MAX_RETRIES = 3

# Module 12 — Campaigns
CAMPAIGN_TYPES = ["Festival", "Offer", "Service Reminder", "Exchange Offer", "Custom"]
CAMPAIGN_STATUSES = ["Draft", "Scheduled", "Running", "Completed", "Cancelled"]  # show "margin due before X days of delivery"
DOC_REQUIREMENTS = {
    "Booking": ["Aadhaar Card"],
    "Finance": ["Aadhaar Card", "PAN Card", "Bank Passbook"],
    "RTO": ["Aadhaar Card", "PAN Card", "RTO Form 20", "RTO Form 21"],
    "Delivered": ["Aadhaar Card", "Invoice"],
}
ROLES = ["super_admin", "admin", "sales_executive"]

# Module-level permission keys — schema-only (future-ready). Current RBAC still rules.
CRM_MODULES = [
    "leads", "followups", "deals", "bookings", "allotments",
    "documents", "delivery", "payments", "finance", "exchange",
    "whatsapp", "campaigns", "automation", "users", "branches",
    "audit_logs", "masters",
]
PERMISSION_ACTIONS = ["view", "create", "edit", "delete"]

# Configurable knobs
DISCOUNT_APPROVAL_THRESHOLD = 5000.0   # ₹ discount above this needs manager approval
FOLLOWUP_MIN_GAP_SECONDS = 60          # prevent duplicate rapid entries
SLA_HOURS_NO_FOLLOWUP = 24             # hours after creation w/o follow-up -> escalate
AT_RISK_MISSED_COUNT = 2               # missed follow-ups to mark lead at risk


# ============================================================
# Auth helpers
# ============================================================

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True,
                        samesite="none", max_age=3600 * 12, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True,
                        samesite="none", max_age=86400 * 7, path="/")


async def get_current_user(request: Request) -> Dict[str, Any]:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("type") != "access":
            raise HTTPException(401, "Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")


def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(403, "Insufficient permissions")
        return user
    return checker


# ============================================================
# Models (Pydantic)
# ============================================================

class LoginIn(BaseModel):
    email: EmailStr
    password: str


class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    phone: Optional[str] = None
    role: str = "sales_executive"
    branch_id: Optional[str] = None
    reporting_manager_id: Optional[str] = None
    joining_date: Optional[str] = None  # YYYY-MM-DD
    permissions: Optional[Dict[str, Dict[str, bool]]] = None  # future-ready, not enforced


class UserUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    branch_id: Optional[str] = None
    reporting_manager_id: Optional[str] = None
    joining_date: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    permissions: Optional[Dict[str, Dict[str, bool]]] = None


class BranchIn(BaseModel):
    name: str
    code: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    assigned_admin_id: Optional[str] = None
    is_active: Optional[bool] = True
    allow_login_when_inactive: Optional[bool] = True


class BranchUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    assigned_admin_id: Optional[str] = None
    is_active: Optional[bool] = None
    allow_login_when_inactive: Optional[bool] = None


class BrandIn(BaseModel):
    name: str


class ModelIn(BaseModel):
    name: str
    brand_id: str


class VariantIn(BaseModel):
    name: str
    model_id: str


class ColorIn(BaseModel):
    name: str
    hex: Optional[str] = None


class ExchangeInfo(BaseModel):
    old_model: Optional[str] = None
    registration_number: Optional[str] = None
    model_year: Optional[int] = None
    tyre_condition: Optional[str] = None
    battery_condition: Optional[str] = None
    body_condition: Optional[str] = None
    self_start: Optional[bool] = None
    finance_on_rc: Optional[bool] = None
    expected_price: Optional[float] = None
    offered_price: Optional[float] = None
    final_value: Optional[float] = None
    broker_value: Optional[float] = None
    broker_remarks: Optional[str] = None
    photos: List[str] = []
    rc_url: Optional[str] = None
    notes: Optional[str] = None


class DealInfo(BaseModel):
    customer_expected_price: Optional[float] = None
    offered_price: Optional[float] = None
    discount: Optional[float] = None
    interest_level: Optional[str] = None  # Hot/Warm/Cold
    ex_showroom_price: Optional[float] = None
    final_deal_price: Optional[float] = None
    deal_status: Optional[str] = None  # Draft/Negotiation/Approval Pending/Approved/Rejected/Converted
    approval_required: Optional[bool] = None
    approval_status: Optional[str] = None  # Pending / Approved / Rejected
    approved_by: Optional[str] = None
    approved_by_name: Optional[str] = None
    approved_at: Optional[str] = None
    approval_remarks: Optional[str] = None


class FinanceInfo(BaseModel):
    finance_company: Optional[str] = None
    down_payment: Optional[float] = None
    emi: Optional[float] = None
    tenure: Optional[int] = None


class RegistrationInfo(BaseModel):
    status: Optional[str] = None  # Pending / Allotted / Plate Fitted / Done
    rto_office: Optional[str] = None
    number_allotted: Optional[str] = None  # e.g. GJ-15-AB-1234
    number_allotted_date: Optional[str] = None  # YYYY-MM-DD
    plate_fitted: Optional[bool] = None
    plate_fitted_date: Optional[str] = None
    fitness_certificate_url: Optional[str] = None
    registration_copy_url: Optional[str] = None
    notes: Optional[str] = None


class LeadCreate(BaseModel):
    customer_name: str
    phone: str
    alt_phone: Optional[str] = None
    birthdate: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    source: Optional[str] = "Walk-in"
    branch_id: Optional[str] = None
    priority: str = "Warm"
    assigned_to: Optional[str] = None  # user id, if None auto-assign
    brand_id: Optional[str] = None
    model_id: Optional[str] = None
    variant_id: Optional[str] = None
    color_id: Optional[str] = None
    vehicle_type: Optional[str] = None  # Bike / Scooty
    test_ride_done: Optional[bool] = None
    purchase_type: Optional[str] = "New Purchase"  # or "Exchange Vehicle"
    customer_type: Optional[str] = None  # "Instant Buyer" | "Token Finance Buyer" | "Just Inquiry"
    exchange: Optional[ExchangeInfo] = None
    deal: Optional[DealInfo] = None
    payment_mode: Optional[str] = None
    finance: Optional[FinanceInfo] = None
    next_followup_date: Optional[str] = None
    next_followup_type: Optional[str] = None
    notes: Optional[str] = None


class LeadUpdate(BaseModel):
    customer_name: Optional[str] = None
    phone: Optional[str] = None
    alt_phone: Optional[str] = None
    birthdate: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    branch_id: Optional[str] = None
    assigned_to: Optional[str] = None
    source: Optional[str] = None
    priority: Optional[str] = None
    brand_id: Optional[str] = None
    model_id: Optional[str] = None
    variant_id: Optional[str] = None
    color_id: Optional[str] = None
    vehicle_type: Optional[str] = None
    test_ride_done: Optional[bool] = None
    purchase_type: Optional[str] = None
    customer_type: Optional[str] = None
    exchange: Optional[ExchangeInfo] = None
    deal: Optional[DealInfo] = None
    payment_mode: Optional[str] = None
    finance: Optional[FinanceInfo] = None
    registration: Optional[RegistrationInfo] = None
    next_followup_date: Optional[str] = None
    next_followup_time: Optional[str] = None
    next_followup_type: Optional[str] = None
    notes: Optional[str] = None
    stage: Optional[str] = None  # admin/super_admin override only (skips form-gating)


class StageChange(BaseModel):
    stage: str
    lost_reason: Optional[str] = None
    lost_reason_text: Optional[str] = None


class FollowupIn(BaseModel):
    type: str
    notes: str
    scheduled_date: Optional[str] = None  # YYYY-MM-DD
    scheduled_time: Optional[str] = None  # HH:MM
    done: bool = True  # marks this follow-up as completed (call/visit done)
    call_status: Optional[str] = None
    customer_response: Optional[str] = None
    outcome_tag: Optional[str] = None
    lead_temperature: Optional[str] = None  # Hot/Warm/Cold
    loss_reason: Optional[str] = None
    loss_reason_text: Optional[str] = None
    call_duration: Optional[int] = None  # seconds
    call_recording_url: Optional[str] = None  # object storage URL
    call_recording_filename: Optional[str] = None


class DealApproveIn(BaseModel):
    approve: bool
    remarks: Optional[str] = None


class DealLossIn(BaseModel):
    reason: str
    text: Optional[str] = None


class BookingIn(BaseModel):
    booking_date: Optional[str] = None  # YYYY-MM-DD
    expected_delivery_date: str
    booking_amount: float
    brand_id: Optional[str] = None
    model_id: Optional[str] = None
    variant_id: Optional[str] = None
    color_id: Optional[str] = None
    finance_company: Optional[str] = None
    down_payment: Optional[float] = None
    emi: Optional[float] = None
    loan_status: Optional[str] = None
    exchange_final_value: Optional[float] = None
    notes: Optional[str] = None
    payment_type: Optional[str] = "Token"  # "Token" or "Full"
    inventory_id: Optional[str] = None     # picked from /inventory
    chassis_number: Optional[str] = None   # locked chassis from inventory


class BookingUpdate(BaseModel):
    booking_date: Optional[str] = None
    expected_delivery_date: Optional[str] = None
    booking_amount: Optional[float] = None
    brand_id: Optional[str] = None
    model_id: Optional[str] = None
    variant_id: Optional[str] = None
    color_id: Optional[str] = None
    finance_company: Optional[str] = None
    down_payment: Optional[float] = None
    emi: Optional[float] = None
    loan_status: Optional[str] = None
    exchange_final_value: Optional[float] = None
    notes: Optional[str] = None
    payment_type: Optional[str] = None
    inventory_id: Optional[str] = None
    chassis_number: Optional[str] = None
    status: Optional[str] = None


class PaymentIn(BaseModel):
    amount: float
    date: Optional[str] = None  # YYYY-MM-DD
    mode: str  # Cash/UPI/Bank Transfer/Finance/Cheque
    payment_type: Optional[str] = "Booking"  # Booking/Margin/Final/Other
    notes: Optional[str] = None


class FinanceCaseIn(BaseModel):
    finance_company: str
    downpayment_amount: Optional[float] = None
    emi: Optional[float] = None
    tenure: Optional[int] = None
    assigned_to: Optional[str] = None


class FinanceCaseUpdate(BaseModel):
    finance_company: Optional[str] = None
    downpayment_amount: Optional[float] = None
    emi: Optional[float] = None
    tenure: Optional[int] = None
    assigned_to: Optional[str] = None
    status: Optional[str] = None
    rejection_reason: Optional[str] = None
    eligibility_notes: Optional[str] = None
    downpayment_received: Optional[bool] = None


class ExchangeValuationIn(BaseModel):
    source: str  # "broker" / "internal" / "online"
    value: float
    remarks: Optional[str] = None


# ---- Module 11: WhatsApp ----

class WATemplateIn(BaseModel):
    name: str
    category: Optional[str] = "general"
    message_type: str = "text"  # text/image/pdf/document
    body: str
    media_url: Optional[str] = None
    active: bool = True


class WATemplateUpdate(BaseModel):
    name: Optional[str] = None
    category: Optional[str] = None
    message_type: Optional[str] = None
    body: Optional[str] = None
    media_url: Optional[str] = None
    active: Optional[bool] = None


class AutomationRuleIn(BaseModel):
    name: str
    event: str  # in WA_EVENTS
    conditions: Dict[str, Any] = {}  # {source, priority, purchase_type, payment_mode, branch_id, stage, brand_id}
    template_id: str
    delay_minutes: int = 0
    active: bool = True


class AutomationRuleUpdate(BaseModel):
    name: Optional[str] = None
    event: Optional[str] = None
    conditions: Optional[Dict[str, Any]] = None
    template_id: Optional[str] = None
    delay_minutes: Optional[int] = None
    active: Optional[bool] = None


class ManualMessageIn(BaseModel):
    template_id: Optional[str] = None
    content: Optional[str] = None
    message_type: str = "text"
    media_url: Optional[str] = None
    variables: Dict[str, str] = {}


class InboundMessageIn(BaseModel):
    content: str
    reply_tag: Optional[str] = None


class MessageMarkIn(BaseModel):
    status: str  # SENT / DELIVERED / READ / FAILED


# ---- Module 12: Campaigns ----

class CampaignTarget(BaseModel):
    stages: List[str] = []
    priorities: List[str] = []
    sources: List[str] = []
    branch_ids: List[str] = []
    purchase_types: List[str] = []
    audience: str = "leads"  # leads / past_buyers / all


class CampaignIn(BaseModel):
    name: str
    campaign_type: str = "Custom"
    template_id: Optional[str] = None
    message_type: str = "text"
    content: Optional[str] = None
    media_url: Optional[str] = None
    scheduled_at: Optional[str] = None  # ISO
    target: CampaignTarget = CampaignTarget()


class CampaignUpdate(BaseModel):
    name: Optional[str] = None
    campaign_type: Optional[str] = None
    template_id: Optional[str] = None
    message_type: Optional[str] = None
    content: Optional[str] = None
    media_url: Optional[str] = None
    scheduled_at: Optional[str] = None
    target: Optional[CampaignTarget] = None
    status: Optional[str] = None


class AllotmentIn(BaseModel):
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None


class AllotmentUpdate(BaseModel):
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None
    status: Optional[str] = None


# ---- Module 7: Documents ----

class ExtractedData(BaseModel):
    document_number: Optional[str] = None
    name: Optional[str] = None
    address: Optional[str] = None
    chassis_number: Optional[str] = None
    engine_number: Optional[str] = None
    vehicle_model: Optional[str] = None
    variant: Optional[str] = None
    confidence_score: Optional[float] = None


class DocumentUpdate(BaseModel):
    doc_type: Optional[str] = None
    doc_number: Optional[str] = None
    extracted: Optional[ExtractedData] = None
    notes: Optional[str] = None


class DocumentReject(BaseModel):
    reason: str


# ---- Module 8: Delivery ----

class AccessoryItem(BaseModel):
    name: str
    quantity: int = 1
    value: float = 0.0


class DeliveryChecklist(BaseModel):
    payment_completed: bool = False
    documents_verified: bool = False
    vehicle_ready: bool = False
    accessories_ready: bool = False


class DeliveryIn(BaseModel):
    delivery_date: str  # YYYY-MM-DD
    time_slot: Optional[str] = None  # "10:00-12:00"
    delivered_by: Optional[str] = None
    instant_bypass: Optional[bool] = False
    bypass_reason: Optional[str] = None
    notes: Optional[str] = None


class DeliveryUpdate(BaseModel):
    delivery_date: Optional[str] = None
    time_slot: Optional[str] = None
    delivered_by: Optional[str] = None
    checklist: Optional[DeliveryChecklist] = None
    accessories: Optional[List[AccessoryItem]] = None
    notes: Optional[str] = None
    status: Optional[str] = None


# ============================================================
# Utility
# ============================================================

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def public_user(u: dict) -> dict:
    u = dict(u)
    u.pop("_id", None)
    u.pop("password_hash", None)
    return u


async def add_timeline(lead_id: str, event: str, actor: dict, meta: Optional[dict] = None):
    await db.timeline.insert_one({
        "id": str(uuid.uuid4()),
        "lead_id": lead_id,
        "event": event,
        "meta": meta or {},
        "actor_id": actor.get("id"),
        "actor_name": actor.get("name"),
        "created_at": now_iso(),
    })


async def log_audit(
    actor: Optional[dict],
    action: str,
    *,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    meta: Optional[dict] = None,
    status: str = "success",
):
    """Append-only audit log. Never raises — audit MUST NOT break the request."""
    try:
        await db.audit_logs.insert_one({
            "id": str(uuid.uuid4()),
            "actor_id": (actor or {}).get("id"),
            "actor_name": (actor or {}).get("name"),
            "actor_role": (actor or {}).get("role"),
            "action": action,
            "entity_type": entity_type,
            "entity_id": entity_id,
            "branch_id": branch_id or (actor or {}).get("branch_id"),
            "status": status,
            "meta": meta or {},
            "created_at": now_iso(),
        })
    except Exception as e:
        logger.warning(f"audit log failed: {e}")


async def can_access_lead(user: dict, lead: dict) -> bool:
    if user["role"] == "super_admin":
        return True
    if user["role"] == "admin":
        return lead.get("branch_id") == user.get("branch_id")
    return lead.get("assigned_to") == user["id"]


# ============================================================
# Object storage
# ============================================================

_storage_key: Optional[str] = None


def init_storage() -> Optional[str]:
    global _storage_key
    if _storage_key:
        return _storage_key
    if not EMERGENT_LLM_KEY:
        return None
    try:
        resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_LLM_KEY}, timeout=30)
        resp.raise_for_status()
        _storage_key = resp.json()["storage_key"]
        return _storage_key
    except Exception as e:
        logger.error(f"Storage init failed: {e}")
        return None


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    if not key:
        raise HTTPException(500, "Storage not available")
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120,
    )
    resp.raise_for_status()
    return resp.json()


def get_object(path: str):
    key = init_storage()
    if not key:
        raise HTTPException(500, "Storage not available")
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60,
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")


# ============================================================
# Auth endpoints
# ============================================================

@api.post("/auth/login")
async def login(body: LoginIn, response: Response, request: Request):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        await log_audit(None, "login_failed",
                        entity_type="user", meta={"email": email},
                        status="failed")
        raise HTTPException(401, "Invalid email or password")
    if not user.get("is_active", True):
        await log_audit(user, "login_failed",
                        entity_type="user", entity_id=user["id"],
                        meta={"reason": "account_disabled"}, status="failed")
        raise HTTPException(403, "Account disabled")
    # Branch-inactive gate (per-branch override)
    if user.get("branch_id"):
        branch = await db.branches.find_one({"id": user["branch_id"]}, {"_id": 0})
        if branch and branch.get("is_active") is False and not branch.get("allow_login_when_inactive", True):
            await log_audit(user, "login_failed",
                            entity_type="user", entity_id=user["id"],
                            meta={"reason": "branch_inactive", "branch_id": branch["id"]},
                            status="failed")
            raise HTTPException(403, "Your branch is currently inactive")
    access = create_access_token(user["id"], user["email"], user["role"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    await log_audit(user, "login", entity_type="user", entity_id=user["id"])
    return {"user": public_user(user), "access_token": access}


@api.post("/auth/logout")
async def logout(response: Response, request: Request):
    # Best-effort actor extraction (do not fail logout on missing token)
    actor = None
    try:
        actor = await get_current_user(request)  # type: ignore
    except Exception:
        pass
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    if actor:
        await log_audit(actor, "logout", entity_type="user", entity_id=actor.get("id"))
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ============================================================
# Users
# ============================================================

@api.get("/users")
async def list_users(
    role: Optional[str] = None,
    branch_id: Optional[str] = None,
    status: Optional[str] = None,  # "active" | "inactive"
    q: Optional[str] = None,  # name/email/phone search
    user: dict = Depends(get_current_user),
):
    query: Dict[str, Any] = {}
    if user["role"] == "admin":
        query = {"$or": [{"branch_id": user.get("branch_id")}, {"role": "admin"}]}
    elif user["role"] == "sales_executive":
        query = {"id": user["id"]}
    if role:
        if role not in ROLES:
            raise HTTPException(400, "Invalid role filter")
        query["role"] = role
    if branch_id:
        query["branch_id"] = branch_id
    if status == "active":
        query["is_active"] = True
    elif status == "inactive":
        query["is_active"] = False
    if q:
        qx = {"$regex": q, "$options": "i"}
        query["$and"] = [{"$or": [{"name": qx}, {"email": qx}, {"phone": qx}]}]
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


@api.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(require_roles("super_admin"))):
    if body.role not in ROLES:
        raise HTTPException(400, "Invalid role")
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email already exists")
    if body.phone and await db.users.find_one({"phone": body.phone}):
        raise HTTPException(400, "Phone number already exists")
    if user["role"] == "admin" and body.role == "super_admin":
        raise HTTPException(403, "Cannot create super admin")
    branch_id = body.branch_id
    if user["role"] == "admin":
        branch_id = user.get("branch_id")
    # Validate reporting_manager (must be admin or super_admin)
    if body.reporting_manager_id:
        mgr = await db.users.find_one({"id": body.reporting_manager_id}, {"_id": 0})
        if not mgr or mgr["role"] not in ("admin", "super_admin"):
            raise HTTPException(400, "reporting_manager_id must be an admin or super_admin user")
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hash_password(body.password),
        "name": body.name,
        "phone": body.phone,
        "role": body.role,
        "branch_id": branch_id,
        "reporting_manager_id": body.reporting_manager_id,
        "joining_date": body.joining_date,
        "permissions": body.permissions or {},
        "is_active": True,
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    await log_audit(user, "user_created", entity_type="user", entity_id=doc["id"],
                    meta={"role": body.role, "branch_id": branch_id})
    return public_user(doc)


@api.put("/users/{uid}")
async def update_user(uid: str, body: UserUpdate, user: dict = Depends(require_roles("super_admin"))):
    target = await db.users.find_one({"id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    if user["role"] == "admin" and target.get("branch_id") != user.get("branch_id") and target["id"] != user["id"]:
        raise HTTPException(403, "Cannot edit users outside branch")
    if user["role"] == "admin" and body.role == "super_admin":
        raise HTTPException(403, "Cannot promote to super admin")
    # phone uniqueness
    if body.phone and body.phone != target.get("phone"):
        dup = await db.users.find_one({"phone": body.phone, "id": {"$ne": uid}})
        if dup:
            raise HTTPException(400, "Phone number already in use")
    if body.reporting_manager_id:
        mgr = await db.users.find_one({"id": body.reporting_manager_id}, {"_id": 0})
        if not mgr or mgr["role"] not in ("admin", "super_admin"):
            raise HTTPException(400, "reporting_manager_id must be an admin or super_admin user")
    updates: Dict[str, Any] = {}
    for field in ["name", "phone", "role", "branch_id", "reporting_manager_id",
                  "joining_date", "is_active", "permissions"]:
        v = getattr(body, field)
        if v is not None:
            updates[field] = v
    if body.password:
        updates["password_hash"] = hash_password(body.password)
    if updates:
        await db.users.update_one({"id": uid}, {"$set": updates})
        await log_audit(user, "user_updated", entity_type="user", entity_id=uid,
                        meta={"fields": list(updates.keys())})
    fresh = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    return fresh


@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require_roles("super_admin"))):
    if uid == user["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    target = await db.users.find_one({"id": uid}, {"_id": 0})
    await db.users.delete_one({"id": uid})
    await log_audit(user, "user_deleted", entity_type="user", entity_id=uid,
                    meta={"email": (target or {}).get("email")})
    return {"ok": True}


@api.get("/users/{uid}/performance")
async def user_performance(uid: str, user: dict = Depends(get_current_user)):
    target = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    if not target:
        raise HTTPException(404, "User not found")
    # Visibility: super_admin all; admin for own branch; sales_executive for self only
    if user["role"] == "sales_executive" and uid != user["id"]:
        raise HTTPException(403, "Access denied")
    if user["role"] == "admin" and target.get("branch_id") != user.get("branch_id") and uid != user["id"]:
        raise HTTPException(403, "Access denied")
    leads_total = await db.leads.count_documents({"assigned_to": uid})
    leads_lost = await db.leads.count_documents({"assigned_to": uid, "stage": "Lost"})
    leads_delivered = await db.leads.count_documents({"assigned_to": uid, "stage": {"$in": ["Delivered", "RTO", "Delivered"]}})
    followups_total = await db.followups.count_documents({"created_by": uid})
    pending = await db.leads.count_documents({
        "assigned_to": uid,
        "stage": {"$nin": ["Lost", "Delivered", "RTO", "Delivered"]}
    })
    conv_rate = round((leads_delivered / leads_total * 100), 1) if leads_total else 0.0
    return {
        "user_id": uid,
        "name": target.get("name"),
        "branch_id": target.get("branch_id"),
        "leads_total": leads_total,
        "leads_lost": leads_lost,
        "leads_delivered": leads_delivered,
        "leads_pending": pending,
        "followups_total": followups_total,
        "conversion_rate_pct": conv_rate,
    }


# ============================================================
# Master data
# ============================================================

async def _list_collection(name: str):
    return await db[name].find({}, {"_id": 0}).sort("name", 1).to_list(1000)


# Branches
@api.get("/branches")
async def list_branches(
    is_active: Optional[bool] = None,
    user: dict = Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if is_active is not None:
        q["is_active"] = is_active
    items = await db.branches.find(q, {"_id": 0}).sort("name", 1).to_list(1000)
    return items


@api.get("/branches/{bid}")
async def get_branch(bid: str, user: dict = Depends(get_current_user)):
    b = await db.branches.find_one({"id": bid}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Branch not found")
    return b


@api.post("/branches")
async def create_branch(body: BranchIn, user: dict = Depends(require_roles("super_admin"))):
    if body.code:
        if await db.branches.find_one({"code": body.code}):
            raise HTTPException(400, "Branch code already exists")
    if body.assigned_admin_id:
        mgr = await db.users.find_one({"id": body.assigned_admin_id}, {"_id": 0})
        if not mgr or mgr["role"] not in ("admin", "super_admin"):
            raise HTTPException(400, "assigned_admin_id must be an admin or super_admin user")
    doc = {
        "id": str(uuid.uuid4()),
        "name": body.name,
        "code": body.code,
        "city": body.city,
        "address": body.address,
        "assigned_admin_id": body.assigned_admin_id,
        "is_active": True if body.is_active is None else body.is_active,
        "allow_login_when_inactive": True if body.allow_login_when_inactive is None else body.allow_login_when_inactive,
        "created_at": now_iso(),
    }
    await db.branches.insert_one(doc)
    await log_audit(user, "branch_created", entity_type="branch", entity_id=doc["id"],
                    meta={"name": doc["name"], "code": doc["code"]})
    return {k: v for k, v in doc.items() if k != "_id"}


@api.put("/branches/{bid}")
async def update_branch(bid: str, body: BranchUpdate, user: dict = Depends(require_roles("super_admin"))):
    target = await db.branches.find_one({"id": bid}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Branch not found")
    if body.code and body.code != target.get("code"):
        dup = await db.branches.find_one({"code": body.code, "id": {"$ne": bid}})
        if dup:
            raise HTTPException(400, "Branch code already in use")
    if body.assigned_admin_id:
        mgr = await db.users.find_one({"id": body.assigned_admin_id}, {"_id": 0})
        if not mgr or mgr["role"] not in ("admin", "super_admin"):
            raise HTTPException(400, "assigned_admin_id must be an admin or super_admin user")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if updates:
        await db.branches.update_one({"id": bid}, {"$set": updates})
        await log_audit(user, "branch_updated", entity_type="branch", entity_id=bid,
                        meta={"fields": list(updates.keys())})
    return await db.branches.find_one({"id": bid}, {"_id": 0})


@api.delete("/branches/{bid}")
async def delete_branch(bid: str, user: dict = Depends(require_roles("super_admin"))):
    # Safety: block deletion if users or leads reference this branch
    uc = await db.users.count_documents({"branch_id": bid})
    lc = await db.leads.count_documents({"branch_id": bid})
    if uc or lc:
        raise HTTPException(400, f"Cannot delete — {uc} users and {lc} leads are linked. Deactivate instead.")
    await db.branches.delete_one({"id": bid})
    await log_audit(user, "branch_deleted", entity_type="branch", entity_id=bid)
    return {"ok": True}


@api.get("/branches/{bid}/performance")
async def branch_performance(bid: str, user: dict = Depends(get_current_user)):
    branch = await db.branches.find_one({"id": bid}, {"_id": 0})
    if not branch:
        raise HTTPException(404, "Branch not found")
    if user["role"] == "admin" and bid != user.get("branch_id"):
        raise HTTPException(403, "Access denied")
    if user["role"] == "sales_executive" and bid != user.get("branch_id"):
        raise HTTPException(403, "Access denied")
    leads_total = await db.leads.count_documents({"branch_id": bid})
    leads_lost = await db.leads.count_documents({"branch_id": bid, "stage": "Lost"})
    leads_delivered = await db.leads.count_documents({
        "branch_id": bid, "stage": {"$in": ["Delivered", "RTO", "Delivered"]}
    })
    users_count = await db.users.count_documents({"branch_id": bid, "is_active": True})
    # Revenue: sum of bookings.final_deal_price for delivered bookings in this branch
    pipeline = [
        {"$match": {"branch_id": bid, "status": "Delivered"}},
        {"$group": {"_id": None, "total": {"$sum": "$final_deal_price"}}},
    ]
    agg = await db.bookings.aggregate(pipeline).to_list(1)
    revenue = float(agg[0]["total"]) if agg else 0.0
    conv_rate = round((leads_delivered / leads_total * 100), 1) if leads_total else 0.0
    return {
        "branch_id": bid,
        "name": branch.get("name"),
        "code": branch.get("code"),
        "is_active": branch.get("is_active", True),
        "leads_total": leads_total,
        "leads_lost": leads_lost,
        "leads_delivered": leads_delivered,
        "conversion_rate_pct": conv_rate,
        "active_users": users_count,
        "revenue": revenue,
    }


@api.get("/branches-compare")
async def branches_compare(user: dict = Depends(require_roles("super_admin"))):
    branches = await db.branches.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    out = []
    for b in branches:
        leads_total = await db.leads.count_documents({"branch_id": b["id"]})
        leads_lost = await db.leads.count_documents({"branch_id": b["id"], "stage": "Lost"})
        leads_delivered = await db.leads.count_documents({
            "branch_id": b["id"], "stage": {"$in": ["Delivered", "RTO", "Delivered"]}
        })
        pipeline = [
            {"$match": {"branch_id": b["id"], "status": "Delivered"}},
            {"$group": {"_id": None, "total": {"$sum": "$final_deal_price"}}},
        ]
        agg = await db.bookings.aggregate(pipeline).to_list(1)
        revenue = float(agg[0]["total"]) if agg else 0.0
        conv_rate = round((leads_delivered / leads_total * 100), 1) if leads_total else 0.0
        out.append({
            "branch_id": b["id"],
            "name": b["name"],
            "code": b.get("code"),
            "is_active": b.get("is_active", True),
            "leads_total": leads_total,
            "leads_lost": leads_lost,
            "leads_delivered": leads_delivered,
            "conversion_rate_pct": conv_rate,
            "revenue": revenue,
        })
    return out


# ============================================================
# Audit Logs
# ============================================================

@api.get("/audit-logs")
async def list_audit_logs(
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    since: Optional[str] = None,  # ISO datetime
    until: Optional[str] = None,
    limit: int = 200,
    user: dict = Depends(require_roles("super_admin")),
):
    q: Dict[str, Any] = {}
    # Super-admin only — admins no longer have audit access
    if user["role"] == "admin":
        q["branch_id"] = user.get("branch_id")
    if user_id:
        q["actor_id"] = user_id
    if action:
        q["action"] = action
    if entity_type:
        q["entity_type"] = entity_type
    if entity_id:
        q["entity_id"] = entity_id
    if since or until:
        created: Dict[str, Any] = {}
        if since:
            created["$gte"] = since
        if until:
            created["$lte"] = until
        q["created_at"] = created
    limit = max(1, min(limit, 1000))
    items = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return items


@api.get("/permissions/modules")
async def list_permission_modules(user: dict = Depends(require_roles("super_admin", "admin"))):
    """Return the module & action catalog for the permissions UI (future-ready)."""
    return {"modules": CRM_MODULES, "actions": PERMISSION_ACTIONS}


# Brands
@api.get("/brands")
async def list_brands(user: dict = Depends(get_current_user)):
    return await _list_collection("brands")


@api.post("/brands")
async def create_brand(body: BrandIn, user: dict = Depends(require_roles("super_admin"))):
    doc = {"id": str(uuid.uuid4()), "name": body.name, "created_at": now_iso()}
    await db.brands.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.delete("/brands/{bid}")
async def delete_brand(bid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.brands.delete_one({"id": bid})
    return {"ok": True}


# Models
@api.get("/models")
async def list_models(brand_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = {}
    if brand_id:
        q["brand_id"] = brand_id
    return await db.vehicle_models.find(q, {"_id": 0}).sort("name", 1).to_list(1000)


@api.post("/models")
async def create_model(body: ModelIn, user: dict = Depends(require_roles("super_admin"))):
    doc = {"id": str(uuid.uuid4()), "name": body.name, "brand_id": body.brand_id, "created_at": now_iso()}
    await db.vehicle_models.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.delete("/models/{mid}")
async def delete_model(mid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.vehicle_models.delete_one({"id": mid})
    return {"ok": True}


# Variants
@api.get("/variants")
async def list_variants(model_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = {}
    if model_id:
        q["model_id"] = model_id
    return await db.variants.find(q, {"_id": 0}).sort("name", 1).to_list(1000)


@api.post("/variants")
async def create_variant(body: VariantIn, user: dict = Depends(require_roles("super_admin"))):
    doc = {"id": str(uuid.uuid4()), "name": body.name, "model_id": body.model_id, "created_at": now_iso()}
    await db.variants.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.delete("/variants/{vid}")
async def delete_variant(vid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.variants.delete_one({"id": vid})
    return {"ok": True}


# Colors
@api.get("/colors")
async def list_colors(user: dict = Depends(get_current_user)):
    return await _list_collection("colors")


@api.post("/colors")
async def create_color(body: ColorIn, user: dict = Depends(require_roles("super_admin"))):
    doc = {"id": str(uuid.uuid4()), "name": body.name, "hex": body.hex, "created_at": now_iso()}
    await db.colors.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


@api.delete("/colors/{cid}")
async def delete_color(cid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.colors.delete_one({"id": cid})
    return {"ok": True}


# ============================================================
# Leads
# ============================================================

async def round_robin_assign(branch_id: str) -> Optional[str]:
    execs = await db.users.find(
        {"role": "sales_executive", "branch_id": branch_id, "is_active": True},
        {"_id": 0, "password_hash": 0}
    ).sort("created_at", 1).to_list(1000)
    if not execs:
        return None
    counter = await db.rr_counters.find_one({"branch_id": branch_id})
    idx = ((counter or {}).get("idx", -1) + 1) % len(execs)
    await db.rr_counters.update_one(
        {"branch_id": branch_id},
        {"$set": {"idx": idx}},
        upsert=True,
    )
    return execs[idx]["id"]


@api.get("/leads")
async def list_leads(
    source: Optional[str] = None,
    stage: Optional[str] = None,
    assigned_to: Optional[str] = None,
    branch_id: Optional[str] = None,
    priority: Optional[str] = None,
    followup_due_today: Optional[bool] = False,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if user["role"] == "sales_executive":
        q["assigned_to"] = user["id"]
    elif user["role"] == "admin":
        q["branch_id"] = user.get("branch_id")
    if source: q["source"] = source
    if stage: q["stage"] = stage
    if assigned_to: q["assigned_to"] = assigned_to
    if branch_id and user["role"] == "super_admin": q["branch_id"] = branch_id
    if priority: q["priority"] = priority
    if followup_due_today:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        q["next_followup_date"] = today
    if search:
        q["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
        ]
    leads = await db.leads.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return leads


@api.post("/leads")
async def create_lead(body: LeadCreate, user: dict = Depends(get_current_user)):
    if body.source and body.source not in LEAD_SOURCES:
        raise HTTPException(400, "Invalid source")
    if body.priority not in PRIORITIES:
        raise HTTPException(400, "Invalid priority")
    # Auto-branch assignment: sales_executive always uses own branch; admin uses own branch
    if user["role"] == "sales_executive":
        if not user.get("branch_id"):
            raise HTTPException(400, "Your user is not mapped to a branch — contact admin")
        body.branch_id = user["branch_id"]
    elif user["role"] == "admin":
        if user.get("branch_id"):
            body.branch_id = user["branch_id"]
    # Super admin: if no branch given, fall back to first active branch
    if not body.branch_id and user["role"] == "super_admin":
        any_branch = await db.branches.find_one({"is_active": {"$ne": False}}, {"_id": 0})
        if any_branch:
            body.branch_id = any_branch["id"]
    branch = await db.branches.find_one({"id": body.branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(400, "Invalid branch — please contact admin")
    if branch.get("is_active") is False:
        raise HTTPException(403, "Branch is inactive — cannot create new leads")
    assigned_to = body.assigned_to
    if user["role"] == "sales_executive":
        assigned_to = user["id"]
    elif not assigned_to:
        assigned_to = await round_robin_assign(body.branch_id)

    lead_id = str(uuid.uuid4())
    doc = body.model_dump()
    # Validate customer_type
    ct = (body.customer_type or "").strip()
    if ct and ct not in CUSTOMER_TYPES:
        raise HTTPException(400, f"Invalid customer_type. Use one of: {CUSTOMER_TYPES}")
    # Default initial stage based on customer type
    initial_stage = "Inquiry"
    if ct == "Just Inquiry":
        initial_stage = "Follow-up"
    doc.update({
        "id": lead_id,
        "assigned_to": assigned_to,
        "stage": initial_stage,
        "customer_type": ct or None,
        "created_by": user["id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "followup_count": 0,
        "documents": [],
        "lost_reason": None,
        "lost_reason_text": None,
    })
    await db.leads.insert_one(doc)
    await add_timeline(lead_id, "Lead Created", user, {"source": body.source})
    if assigned_to:
        await add_timeline(lead_id, "Lead Assigned", user, {"assigned_to": assigned_to})
    try:
        await fire_event("inquiry_created", doc, user)
    except Exception as _e:
        logger.warning(f"inquiry_created fire_event failed: {_e}")
    await log_audit(user, "lead_created", entity_type="lead", entity_id=lead_id,
                    branch_id=body.branch_id,
                    meta={"source": body.source, "priority": body.priority})
    return await db.leads.find_one({"id": lead_id}, {"_id": 0})


@api.get("/leads/{lid}")
async def get_lead(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return lead


@api.put("/leads/{lid}")
async def update_lead(lid: str, body: LeadUpdate, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")

    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "assigned_to" in updates and user["role"] == "sales_executive":
        del updates["assigned_to"]

    # Validate customer_type whitelist on update too
    if "customer_type" in updates:
        ct = (updates["customer_type"] or "").strip()
        if ct and ct not in CUSTOMER_TYPES:
            raise HTTPException(400, f"Invalid customer_type. Use one of: {CUSTOMER_TYPES}")
        updates["customer_type"] = ct or None

    # Stage override (admin/super_admin only) — skips form-gating, useful for fixing mistakes
    if "stage" in updates:
        if user["role"] == "sales_executive":
            raise HTTPException(403, "Sales executives cannot override stage; use Change Stage flow")
        new_stage = updates["stage"]
        if new_stage not in STAGES:
            # try alias
            new_stage = STAGE_ALIAS.get(new_stage, new_stage)
            if new_stage not in STAGES:
                raise HTTPException(400, f"Invalid stage. Use one of: {STAGES}")
        updates["stage"] = new_stage

    # Conditional cleanup — if switching Exchange → New Purchase, wipe exchange docs/photos
    # (Keep identity_docs intact since Aadhaar is common to both.)
    if updates.get("purchase_type") == "New Purchase" and (lead.get("purchase_type") or "") == "Exchange Vehicle":
        updates["exchange"] = None

    # Negotiation history: log if deal fields changed
    if "deal" in updates:
        old_deal = lead.get("deal") or {}
        new_deal = updates["deal"] or {}
        changed = {}
        for k in ["customer_expected_price", "offered_price", "discount",
                  "final_deal_price", "interest_level", "ex_showroom_price"]:
            if new_deal.get(k) != old_deal.get(k):
                changed[k] = {"from": old_deal.get(k), "to": new_deal.get(k)}
        if changed:
            await db.negotiation_history.insert_one({
                "id": str(uuid.uuid4()),
                "lead_id": lid,
                "changed_by": user["id"],
                "changed_by_name": user["name"],
                "changes": changed,
                "note": (updates.get("notes") or None),
                "created_at": now_iso(),
            })

        # Auto-flag approval required when discount exceeds threshold
        discount = new_deal.get("discount") or 0
        if discount and discount >= DISCOUNT_APPROVAL_THRESHOLD:
            new_deal["approval_required"] = True
            if not new_deal.get("approval_status"):
                new_deal["approval_status"] = "Pending"
            if not new_deal.get("deal_status"):
                new_deal["deal_status"] = "Approval Pending"
        updates["deal"] = new_deal

    if updates:
        updates["updated_at"] = now_iso()
        await db.leads.update_one({"id": lid}, {"$set": updates})
        # If exchange.final_value changed, recompute any active booking
        if "exchange" in updates:
            active_b = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}}, {"_id": 0})
            if active_b:
                await _recompute_booking(active_b)
        await add_timeline(lid, "Lead Updated", user, {"fields": list(updates.keys())})
        if "assigned_to" in updates:
            await add_timeline(lid, "Lead Assigned", user, {"assigned_to": updates["assigned_to"]})
        await log_audit(user, "lead_updated", entity_type="lead", entity_id=lid,
                        branch_id=lead.get("branch_id"),
                        meta={"fields": list(updates.keys())})
    return await db.leads.find_one({"id": lid}, {"_id": 0})


@api.delete("/leads/{lid}")
async def delete_lead(lid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    """Permanent cascade delete of a lead and ALL associated data.
    - super_admin: any branch
    - admin: only own branch
    - sales_executive: forbidden
    """
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if user["role"] == "admin" and lead.get("branch_id") != user.get("branch_id"):
        raise HTTPException(403, "Cannot delete leads outside your branch")

    # Release any inventory chassis booked by this lead
    await db.inventory.update_many(
        {"booked_by_lead": lid},
        {"$set": {"status": "Available", "booked_by_lead": None, "booked_at": None}},
    )

    # Cascade delete every collection that may reference lead_id = lid
    related = [
        "followups", "finance_cases", "bookings", "allotments", "deliveries",
        "payments", "wa_messages", "timeline", "negotiation_history",
        "documents", "files", "reminders",
    ]
    deleted_counts = {}
    for coll in related:
        r = await db[coll].delete_many({"lead_id": lid})
        deleted_counts[coll] = r.deleted_count

    # Finally delete the lead itself
    await db.leads.delete_one({"id": lid})

    await log_audit(user, "lead_deleted", entity_type="lead", entity_id=lid,
                    branch_id=lead.get("branch_id"),
                    meta={"customer_name": lead.get("customer_name"),
                          "phone": lead.get("phone"),
                          "stage": lead.get("stage"),
                          "deleted_counts": deleted_counts})
    return {"ok": True, "deleted": deleted_counts}




@api.post("/leads/{lid}/deal/request-approval")
async def deal_request_approval(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    deal = lead.get("deal") or {}
    deal["approval_required"] = True
    deal["approval_status"] = "Pending"
    deal["deal_status"] = "Approval Pending"
    await db.leads.update_one({"id": lid}, {"$set": {"deal": deal, "updated_at": now_iso()}})
    await add_timeline(lid, "Approval Requested", user,
                       {"discount": deal.get("discount"), "final_price": deal.get("final_deal_price")})
    return await db.leads.find_one({"id": lid}, {"_id": 0})


@api.post("/leads/{lid}/deal/approve")
async def deal_approve(lid: str, body: DealApproveIn,
                       user: dict = Depends(require_roles("super_admin", "admin"))):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if user["role"] == "admin" and lead.get("branch_id") != user.get("branch_id"):
        raise HTTPException(403, "Cannot approve deals outside your branch")
    deal = lead.get("deal") or {}
    deal["approval_required"] = True
    deal["approval_status"] = "Approved" if body.approve else "Rejected"
    deal["deal_status"] = "Approved" if body.approve else "Rejected"
    deal["approved_by"] = user["id"]
    deal["approved_by_name"] = user["name"]
    deal["approved_at"] = now_iso()
    deal["approval_remarks"] = body.remarks
    await db.leads.update_one({"id": lid}, {"$set": {"deal": deal, "updated_at": now_iso()}})
    await add_timeline(lid, "Deal " + ("Approved" if body.approve else "Rejected"), user,
                       {"remarks": body.remarks})
    return await db.leads.find_one({"id": lid}, {"_id": 0})


@api.get("/leads/{lid}/negotiations")
async def list_negotiations(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.negotiation_history.find({"lead_id": lid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.post("/leads/{lid}/stage")
async def change_stage(lid: str, body: StageChange, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if body.stage not in STAGES:
        raise HTTPException(400, "Invalid stage")

    # --- STRICT FORM-BASED FUNNEL VALIDATION ---
    # Identity / Ownership document gating — required only when moving to Booking and beyond
    # (Inquiry / Follow-up / Hold / Lost don't require docs)
    DOC_GATED_STAGES = {"Booking", "Delivery", "Allotment", "Feedback"}
    if body.stage in DOC_GATED_STAGES:
        identity = lead.get("identity_docs") or {}
        missing_id = []
        if not (identity.get("aadhaar") or []):
            missing_id.append("Aadhaar")
        if missing_id:
            raise HTTPException(400, f"Upload required KYC documents before moving to {body.stage}: {', '.join(missing_id)}")
        if (lead.get("purchase_type") or "") == "Exchange Vehicle":
            exch_docs = (lead.get("exchange") or {}).get("documents") or {}
            missing_ex = []
            rc_combined = (exch_docs.get("rc") or []) + (exch_docs.get("rc_front") or []) + (exch_docs.get("rc_back") or [])
            if not rc_combined:
                missing_ex.append("RC Book")
            if not (exch_docs.get("front_photo") or []):
                missing_ex.append("Vehicle Front Photo")
            if not (exch_docs.get("back_photo") or []):
                missing_ex.append("Vehicle Back Photo")
            if missing_ex:
                raise HTTPException(400, f"Exchange Vehicle requires: {', '.join(missing_ex)} before moving to {body.stage}")

    # Inquiry → Follow-up: Name + Phone + Vehicle (brand OR model) required
    deal = lead.get("deal") or {}
    if body.stage == "Follow-up":
        if not (lead.get("customer_name") and lead.get("phone")):
            raise HTTPException(400, "Customer name and phone are required before Follow-up")
        if not (lead.get("brand_id") or lead.get("model_id")):
            raise HTTPException(400, "Select a vehicle (brand/model) before Follow-up")

    # Follow-up → Interest: at least 1 successful follow-up (Connected + Interested)
    if body.stage == "Follow-up":
        interested = await db.followups.count_documents({
            "lead_id": lid,
            "call_status": "Connected",
            "customer_response": "Interested",
        })
        if interested == 0:
            raise HTTPException(400, "Need at least one Connected follow-up marked 'Interested' before Interest stage")

    # Interest → Deal: brand, model, budget (customer_expected_price) required
    if body.stage == "Booking":
        if not lead.get("brand_id"):
            raise HTTPException(400, "Brand is required before Deal")
        if not lead.get("model_id"):
            raise HTTPException(400, "Model is required before Deal")
        if not deal.get("customer_expected_price"):
            raise HTTPException(400, "Customer budget (expected price) is required before Deal")
        # Must have at least one Connected follow-up
        connected = await db.followups.count_documents({"lead_id": lid, "call_status": "Connected"})
        if connected == 0:
            raise HTTPException(400, "Log at least one Connected follow-up before moving to Deal")

    # Deal → Booking: final_deal_price + payment_mode required
    if body.stage == "Booking":
        if not lead.get("payment_mode"):
            raise HTTPException(400, "Set payment mode before Booking")
        if not deal.get("final_deal_price"):
            raise HTTPException(400, "Set Final Deal Price before Booking")
        if deal.get("approval_required") and deal.get("approval_status") != "Approved":
            raise HTTPException(400, "Deal requires manager approval before Booking")

    # Booking → Allotment: booking_amount collected (booking exists with amount > 0)
    if body.stage == "Allotment":
        booking = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}}, {"_id": 0})
        if not booking:
            raise HTTPException(400, "Create a booking with booking amount before Allotment")
        if not booking.get("booking_amount") or float(booking.get("booking_amount", 0)) <= 0:
            raise HTTPException(400, "Booking amount is required before Allotment")

    # Allotment → Delivery: chassis number required
    if body.stage == "Delivered":
        allotment = await db.allotments.find_one({"lead_id": lid}, {"_id": 0})
        if not allotment or not allotment.get("chassis_number"):
            raise HTTPException(400, "Chassis number (allotment) is required before Delivery")

    if body.stage == "RTO":
        # Payment must be fully paid
        booking = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}}, {"_id": 0})
        if booking:
            paid = float(booking.get("total_paid") or 0)
            final = float(booking.get("final_deal_price") or 0)
            if final > 0 and paid < final:
                raise HTTPException(400, f"Full payment (₹{final}) must be completed before Registration. Paid: ₹{paid}")
        # Module 7 enforcement: all required Registration docs must be Verified
        lead_docs = await db.documents.find({"lead_id": lid, "is_latest": True}, {"_id": 0}).to_list(200)
        missing = _stage_doc_requirements("RTO", lead_docs)
        if missing:
            raise HTTPException(400, f"Missing verified documents for Registration: {', '.join(missing)}")
    if body.stage == "Lost":
        if not body.lost_reason:
            raise HTTPException(400, "Lost reason is required")

    upd: Dict[str, Any] = {"stage": body.stage, "updated_at": now_iso()}
    if body.stage == "Lost":
        upd["lost_reason"] = body.lost_reason
        upd["lost_reason_text"] = body.lost_reason_text
    if body.stage == "Booking":
        # mark deal as converted
        deal["deal_status"] = "Converted"
        upd["deal"] = deal
    await db.leads.update_one({"id": lid}, {"$set": upd})
    await add_timeline(lid, "Stage Changed", user,
                       {"from": lead.get("stage"), "to": body.stage,
                        "lost_reason": body.lost_reason})
    fresh = await db.leads.find_one({"id": lid}, {"_id": 0})
    try:
        if body.stage == "Lost":
            await fire_event("lost_reengage", fresh, user)
        else:
            await fire_event("stage_changed", fresh, user,
                             {"from_stage": lead.get("stage"), "to_stage": body.stage})
    except Exception as _e:
        logger.warning(f"stage_changed fire_event failed: {_e}")
    # Audit: special flag for deal closure (conversions)
    action = "deal_closed" if body.stage in ("Delivered", "RTO", "Delivered") else "stage_changed"
    if body.stage == "Lost":
        action = "lead_lost"
    await log_audit(user, action, entity_type="lead", entity_id=lid,
                    branch_id=lead.get("branch_id"),
                    meta={"from": lead.get("stage"), "to": body.stage,
                          "lost_reason": body.lost_reason})
    return fresh


@api.post("/leads/{lid}/assign")
async def assign_lead(lid: str, assigned_to: str = Query(...),
                      user: dict = Depends(require_roles("super_admin", "admin"))):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    target = await db.users.find_one({"id": assigned_to}, {"_id": 0})
    if not target:
        raise HTTPException(400, "Invalid user")
    if user["role"] == "admin" and lead.get("branch_id") != user.get("branch_id"):
        raise HTTPException(403, "Cannot assign leads outside your branch")
    await db.leads.update_one({"id": lid}, {"$set": {"assigned_to": assigned_to, "updated_at": now_iso()}})
    await add_timeline(lid, "Lead Assigned", user, {"assigned_to": assigned_to})
    return await db.leads.find_one({"id": lid}, {"_id": 0})


# ============================================================
# Follow-ups
# ============================================================

@api.get("/leads/{lid}/followups")
async def list_followups(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.followups.find({"lead_id": lid}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.post("/leads/{lid}/followups")
async def add_followup(lid: str, body: FollowupIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if body.type not in FOLLOWUP_TYPES:
        raise HTTPException(400, "Invalid follow-up type")
    if body.call_status and body.call_status not in CALL_STATUSES:
        raise HTTPException(400, "Invalid call status")
    if body.customer_response and body.customer_response not in CUSTOMER_RESPONSES:
        raise HTTPException(400, "Invalid customer response")
    if body.outcome_tag and body.outcome_tag not in OUTCOME_TAGS:
        raise HTTPException(400, "Invalid outcome tag")
    if not body.scheduled_date:
        raise HTTPException(400, "Next follow-up date is required")

    # Duplicate / rapid entry control
    last = await db.followups.find_one({"lead_id": lid},
                                       {"_id": 0}, sort=[("created_at", -1)])
    if last:
        last_ts = datetime.fromisoformat(last["created_at"])
        gap = (datetime.now(timezone.utc) - last_ts).total_seconds()
        if gap < FOLLOWUP_MIN_GAP_SECONDS:
            raise HTTPException(429, f"Please wait {FOLLOWUP_MIN_GAP_SECONDS - int(gap)}s before logging another follow-up")

    sched = body.scheduled_date
    if body.scheduled_time:
        sched_full = f"{body.scheduled_date}T{body.scheduled_time}"
    else:
        sched_full = f"{body.scheduled_date}T10:00"

    doc = {
        "id": str(uuid.uuid4()),
        "lead_id": lid,
        "branch_id": lead.get("branch_id"),
        "assigned_to": lead.get("assigned_to"),
        "type": body.type,
        "notes": body.notes,
        "scheduled_date": sched,
        "scheduled_time": body.scheduled_time,
        "scheduled_at": sched_full,
        "done": body.done,
        "done_at": now_iso() if body.done else None,
        "call_status": body.call_status,
        "customer_response": body.customer_response,
        "outcome_tag": body.outcome_tag,
        "lead_temperature": body.lead_temperature,
        "loss_reason": body.loss_reason,
        "loss_reason_text": body.loss_reason_text,
        "call_duration": body.call_duration,
        "call_recording_url": body.call_recording_url,
        "call_recording_filename": body.call_recording_filename,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
    }
    await db.followups.insert_one(doc)

    # Update lead
    lead_updates: Dict[str, Any] = {
        "next_followup_date": sched,
        "next_followup_time": body.scheduled_time,
        "next_followup_type": body.type,
        "last_followup_at": now_iso(),
        "updated_at": now_iso(),
    }
    if body.lead_temperature and body.lead_temperature in PRIORITIES:
        lead_updates["priority"] = body.lead_temperature

    # Check at-risk (2+ missed)
    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    missed_count = await db.followups.count_documents({
        "lead_id": lid,
        "done": False,
        "scheduled_date": {"$lt": today_iso},
    })
    lead_updates["at_risk"] = missed_count >= AT_RISK_MISSED_COUNT
    lead_updates["missed_followups"] = missed_count

    await db.leads.update_one(
        {"id": lid},
        {"$set": lead_updates, "$inc": {"followup_count": 1}},
    )
    await add_timeline(
        lid, "Follow-up Added", user,
        {"type": body.type, "scheduled_date": sched,
         "call_status": body.call_status, "response": body.customer_response,
         "outcome": body.outcome_tag},
    )
    await log_audit(user, "followup_created", entity_type="followup", entity_id=doc["id"],
                    branch_id=lead.get("branch_id"),
                    meta={"lead_id": lid, "type": body.type,
                          "scheduled_date": sched, "outcome": body.outcome_tag})
    return {k: v for k, v in doc.items() if k != "_id"}


# ============================================================
# Tasks (derived from follow-ups)
# ============================================================

def _task_scope_query(user: dict) -> Dict[str, Any]:
    if user["role"] == "sales_executive":
        return {"assigned_to": user["id"]}
    if user["role"] == "admin":
        return {"branch_id": user.get("branch_id")}
    return {}


@api.get("/tasks")
async def list_tasks(kind: str = Query("today"),
                     user: dict = Depends(get_current_user)):
    """kind: today | missed | upcoming | all"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    base = _task_scope_query(user)
    q: Dict[str, Any] = {**base}
    if kind == "today":
        q.update({"next_followup_date": today})
    elif kind == "missed":
        q.update({"next_followup_date": {"$lt": today},
                  "stage": {"$nin": ["Lost", "Delivered", "RTO", "Delivered"]}})
    elif kind == "upcoming":
        q.update({"next_followup_date": {"$gt": today}})
    elif kind == "at_risk":
        q.update({"at_risk": True})
    leads = await db.leads.find(q, {"_id": 0}).sort("next_followup_date", 1).to_list(500)
    return leads


# ============================================================
# Timeline
# ============================================================

@api.get("/leads/{lid}/timeline")
async def get_timeline(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.timeline.find({"lead_id": lid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


# ============================================================
# Files
# ============================================================

@api.post("/leads/{lid}/documents-quick")
async def upload_document_quick(
    lid: str,
    file: UploadFile = File(...),
    doc_type: str = Form("other"),
    user: dict = Depends(get_current_user),
):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")

    ext = (file.filename or "bin").split(".")[-1].lower() if "." in (file.filename or "") else "bin"
    path = f"{APP_NAME}/leads/{lid}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "application/octet-stream")

    file_id = str(uuid.uuid4())
    frec = {
        "id": file_id,
        "lead_id": lid,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size"),
        "doc_type": doc_type,
        "uploaded_by": user["id"],
        "uploaded_by_name": user["name"],
        "is_deleted": False,
        "created_at": now_iso(),
    }
    await db.files.insert_one(frec)
    await db.leads.update_one(
        {"id": lid},
        {"$push": {"documents": {"id": file_id, "doc_type": doc_type,
                                 "filename": file.filename, "storage_path": result["path"],
                                 "content_type": file.content_type}},
         "$set": {"updated_at": now_iso()}},
    )
    await add_timeline(lid, "Documents Uploaded", user, {"doc_type": doc_type, "filename": file.filename})
    return {k: v for k, v in frec.items() if k != "_id"}


@api.get("/files/{fid}")
async def download_file(fid: str, request: Request,
                        authorization: Optional[str] = Header(None),
                        auth: Optional[str] = Query(None)):
    # Auth via header, cookie, or query token
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
    elif auth:
        token = auth
    else:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except Exception:
        raise HTTPException(401, "Invalid token")

    rec = await db.files.find_one({"id": fid, "is_deleted": False}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "File not found")
    data, ctype = get_object(rec["storage_path"])
    return FastResponse(content=data, media_type=rec.get("content_type") or ctype)


# ============================================================
# Analytics
# ============================================================

@api.get("/analytics/summary")
async def analytics_summary(
    branch_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    base: Dict[str, Any] = {}
    if user["role"] == "sales_executive":
        base["assigned_to"] = user["id"]
    elif user["role"] == "admin":
        base["branch_id"] = user.get("branch_id")
    elif user["role"] == "super_admin" and branch_id:
        base["branch_id"] = branch_id
    if from_date or to_date:
        rng: Dict[str, Any] = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = f"{to_date}T23:59:59"
        base["created_at"] = rng

    total = await db.leads.count_documents(base)

    # per source
    pipeline = [{"$match": base}, {"$group": {"_id": "$source", "count": {"$sum": 1}}}]
    per_source = {}
    async for row in db.leads.aggregate(pipeline):
        per_source[row["_id"] or "Unknown"] = row["count"]

    # per stage
    pipeline2 = [{"$match": base}, {"$group": {"_id": "$stage", "count": {"$sum": 1}}}]
    per_stage = {}
    async for row in db.leads.aggregate(pipeline2):
        per_stage[row["_id"] or "Unknown"] = row["count"]

    converted = await db.leads.count_documents({**base, "stage": {"$in": ["Booking", "Delivered", "RTO", "Delivered"]}})
    lost = await db.leads.count_documents({**base, "stage": "Lost"})

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    due_today = await db.leads.count_documents({**base, "next_followup_date": today})
    missed = await db.leads.count_documents({**base,
                                             "next_followup_date": {"$lt": today},
                                             "stage": {"$nin": ["Lost", "Delivered", "RTO", "Delivered"]}})
    upcoming = await db.leads.count_documents({**base, "next_followup_date": {"$gt": today}})
    at_risk = await db.leads.count_documents({**base, "at_risk": True})

    # Conversion rate
    conversion_rate = round((converted / total) * 100, 1) if total else 0.0

    # Deals in progress
    deals_in_progress = await db.leads.count_documents({**base, "stage": "Booking"})
    # Avg discount
    avg_discount = 0.0
    pipeline3 = [{"$match": {**base, "deal.discount": {"$gt": 0}}},
                 {"$group": {"_id": None, "avg": {"$avg": "$deal.discount"}}}]
    async for row in db.leads.aggregate(pipeline3):
        avg_discount = round(row.get("avg") or 0, 0)

    # Pending approvals
    pending_approvals_q = {**base, "deal.approval_required": True, "deal.approval_status": "Pending"}
    pending_approvals = await db.leads.count_documents(pending_approvals_q)

    return {
        "total_leads": total,
        "per_source": per_source,
        "per_stage": per_stage,
        "converted": converted,
        "lost": lost,
        "followups_due_today": due_today,
        "followups_missed": missed,
        "followups_upcoming": upcoming,
        "at_risk": at_risk,
        "conversion_rate": conversion_rate,
        "deals_in_progress": deals_in_progress,
        "avg_discount": avg_discount,
        "pending_approvals": pending_approvals,
    }


@api.get("/analytics/calendar")
async def analytics_calendar(
    year: int,
    month: int,
    branch_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Month-wise calendar events for Gujarati Calendar popup.
    Returns a map keyed by YYYY-MM-DD with deliveries / followups / upcoming / overdue lists.
    """
    if not (1 <= month <= 12) or year < 2024 or year > 2035:
        raise HTTPException(400, "Invalid year/month")
    # Build month range
    start_iso = f"{year:04d}-{month:02d}-01"
    if month == 12:
        next_month_start = f"{year + 1:04d}-01-01"
    else:
        next_month_start = f"{year:04d}-{month + 1:02d}-01"

    # Scope
    lead_scope: Dict[str, Any] = {}
    if user["role"] == "sales_executive":
        lead_scope["assigned_to"] = user["id"]
    elif user["role"] == "admin":
        lead_scope["branch_id"] = user.get("branch_id")
    elif user["role"] == "super_admin" and branch_id:
        lead_scope["branch_id"] = branch_id

    accessible_lead_ids: Optional[set] = None
    if lead_scope:
        cur = db.leads.find(lead_scope, {"_id": 0, "id": 1})
        accessible_lead_ids = set()
        async for l in cur:
            accessible_lead_ids.add(l["id"])

    def in_scope(lid: str) -> bool:
        return accessible_lead_ids is None or lid in accessible_lead_ids

    days: Dict[str, Dict[str, list]] = {}

    def bucket(date_str: str) -> Dict[str, list]:
        if date_str not in days:
            days[date_str] = {"deliveries": [], "followups": [], "upcoming": [], "overdue": []}
        return days[date_str]

    # 1) Deliveries done — from `deliveries` collection
    del_q: Dict[str, Any] = {"delivery_date": {"$gte": start_iso, "$lt": next_month_start}}
    async for d in db.deliveries.find(del_q, {"_id": 0}):
        if not in_scope(d.get("lead_id", "")):
            continue
        lead = await db.leads.find_one({"id": d.get("lead_id")}, {"_id": 0, "customer_name": 1, "phone": 1, "id": 1})
        bucket(d["delivery_date"])["deliveries"].append({
            "lead_id": d.get("lead_id"),
            "customer_name": (lead or {}).get("customer_name", "—"),
            "phone": (lead or {}).get("phone"),
            "chassis": d.get("chassis_number"),
        })

    # 2) Followups (scheduled/done) — by scheduled_date
    fu_q: Dict[str, Any] = {"scheduled_date": {"$gte": start_iso, "$lt": next_month_start}}
    async for f in db.followups.find(fu_q, {"_id": 0}):
        if not in_scope(f.get("lead_id", "")):
            continue
        lead = await db.leads.find_one({"id": f.get("lead_id")}, {"_id": 0, "customer_name": 1, "phone": 1, "stage": 1})
        bucket(f["scheduled_date"])["followups"].append({
            "lead_id": f.get("lead_id"),
            "customer_name": (lead or {}).get("customer_name", "—"),
            "phone": (lead or {}).get("phone"),
            "stage": (lead or {}).get("stage"),
            "type": f.get("type"),
            "done": bool(f.get("done")),
            "outcome": f.get("outcome_tag"),
        })

    # 3) Upcoming deliveries — bookings with expected_delivery_date in range
    bk_q: Dict[str, Any] = {"expected_delivery_date": {"$gte": start_iso, "$lt": next_month_start}}
    async for bk in db.bookings.find(bk_q, {"_id": 0}):
        if not in_scope(bk.get("lead_id", "")):
            continue
        lead = await db.leads.find_one({"id": bk.get("lead_id")}, {"_id": 0, "customer_name": 1, "phone": 1, "stage": 1})
        bucket(bk["expected_delivery_date"])["upcoming"].append({
            "lead_id": bk.get("lead_id"),
            "customer_name": (lead or {}).get("customer_name", "—"),
            "phone": (lead or {}).get("phone"),
            "stage": (lead or {}).get("stage"),
            "booking_id": bk.get("id"),
            "status": bk.get("status"),
        })

    # 4) Overdue / upcoming from leads.next_followup_date within range
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overdue_stages = ["Lost", "Delivered", "RTO", "Delivered"]
    lead_fu_q: Dict[str, Any] = {
        "next_followup_date": {"$gte": start_iso, "$lt": next_month_start},
        "stage": {"$nin": overdue_stages},
        **lead_scope,
    }
    async for l in db.leads.find(lead_fu_q, {"_id": 0}):
        is_overdue = l.get("next_followup_date") and l["next_followup_date"] < today
        key = "overdue" if is_overdue else "upcoming"
        bucket(l["next_followup_date"])[key].append({
            "lead_id": l.get("id"),
            "customer_name": l.get("customer_name", "—"),
            "phone": l.get("phone"),
            "stage": l.get("stage"),
            "priority": l.get("priority"),
            "followup_time": l.get("next_followup_time"),
            "followup_type": l.get("next_followup_type"),
        })

    return {
        "year": year,
        "month": month,
        "days": days,
    }




@api.get("/analytics/performance")
async def analytics_performance(
    branch_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    """Per sales-executive performance metrics."""
    user_q: Dict[str, Any] = {"role": "sales_executive"}
    lead_base: Dict[str, Any] = {}
    if user["role"] == "admin":
        user_q["branch_id"] = user.get("branch_id")
        lead_base["branch_id"] = user.get("branch_id")
    elif user["role"] == "super_admin" and branch_id:
        user_q["branch_id"] = branch_id
        lead_base["branch_id"] = branch_id
    if from_date or to_date:
        rng: Dict[str, Any] = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = f"{to_date}T23:59:59"
        lead_base["created_at"] = rng

    execs = await db.users.find(user_q, {"_id": 0, "password_hash": 0}).to_list(1000)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    out = []
    for e in execs:
        lq = {**lead_base, "assigned_to": e["id"]}
        total = await db.leads.count_documents(lq)
        converted = await db.leads.count_documents({**lq, "stage": {"$in": ["Booking", "Delivered", "RTO", "Delivered"]}})
        lost = await db.leads.count_documents({**lq, "stage": "Lost"})
        missed = await db.leads.count_documents({**lq, "next_followup_date": {"$lt": today},
                                                  "stage": {"$nin": ["Lost", "Delivered", "RTO", "Delivered"]}})
        at_risk = await db.leads.count_documents({**lq, "at_risk": True})
        fu_total = await db.followups.count_documents({"assigned_to": e["id"]})
        fu_connected = await db.followups.count_documents({"assigned_to": e["id"], "call_status": "Connected"})
        connect_rate = round((fu_connected / fu_total) * 100, 1) if fu_total else 0.0
        conv_rate = round((converted / total) * 100, 1) if total else 0.0
        out.append({
            "user_id": e["id"],
            "name": e["name"],
            "branch_id": e.get("branch_id"),
            "total_leads": total,
            "converted": converted,
            "lost": lost,
            "missed_followups": missed,
            "at_risk": at_risk,
            "followups_logged": fu_total,
            "connect_rate": connect_rate,
            "conversion_rate": conv_rate,
        })
    out.sort(key=lambda x: (-x["converted"], -x["conversion_rate"]))
    return out


# ============================================================
# Bookings (Module 5)
# ============================================================

async def _booking_payment_totals(booking_id: str):
    total = 0.0
    async for p in db.payments.find({"booking_id": booking_id}, {"_id": 0, "amount": 1}):
        total += float(p.get("amount") or 0)
    return total


async def _recompute_booking(booking: dict):
    bid = booking["id"]
    paid = await _booking_payment_totals(bid)
    final_price = float(booking.get("final_deal_price") or 0)
    # Subtract exchange final value from payable
    lead = await db.leads.find_one({"id": booking.get("lead_id")}, {"_id": 0}) if booking.get("lead_id") else None
    exchange_value = float(((lead or {}).get("exchange") or {}).get("final_value") or 0) if lead else 0
    net_payable = max(0.0, final_price - exchange_value)
    pending = net_payable - paid
    status = "Pending" if paid <= 0 else ("Completed" if pending <= 0.01 else "Partial")
    await db.bookings.update_one({"id": bid},
                                 {"$set": {"total_paid": paid, "pending_amount": round(pending, 2),
                                           "net_payable": net_payable,
                                           "exchange_adjustment": exchange_value,
                                           "payment_status": status,
                                           "updated_at": now_iso()}})
    booking["total_paid"] = paid
    booking["pending_amount"] = round(pending, 2)
    booking["net_payable"] = net_payable
    booking["exchange_adjustment"] = exchange_value
    booking["payment_status"] = status
    return booking


@api.post("/leads/{lid}/booking")
async def create_booking(lid: str, body: BookingIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    existing = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}},
                                          {"_id": 0})
    if existing:
        raise HTTPException(400, "Booking already exists for this lead")

    deal = lead.get("deal") or {}
    final_price = deal.get("final_deal_price")
    if not final_price:
        raise HTTPException(400, "Set Final Deal Price on the deal before booking")
    if body.booking_amount <= 0:
        raise HTTPException(400, "Booking amount must be positive")
    if body.booking_amount > float(final_price):
        raise HTTPException(400, "Booking amount cannot exceed final deal price")

    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    booking_date = body.booking_date or today_str
    if body.expected_delivery_date < booking_date:
        raise HTTPException(400, "Expected delivery date must be on or after booking date")

    if body.loan_status and body.loan_status not in LOAN_STATUSES:
        raise HTTPException(400, "Invalid loan status")

    # Stock & chassis lock — if inventory_id provided, validate availability and lock it
    locked_chassis = None
    locked_inv_id = None
    if body.inventory_id:
        inv = await db.inventory.find_one({"id": body.inventory_id}, {"_id": 0})
        if not inv:
            raise HTTPException(400, "Selected inventory item not found")
        if inv.get("status") != "available":
            raise HTTPException(400, f"This vehicle is already {inv.get('status')} — please pick another")
        # Verify model/variant alignment if provided
        locked_chassis = inv.get("chassis_number")
        locked_inv_id = inv.get("id")
    elif body.chassis_number:
        # Allow free-text chassis but check no other booking has it
        dup = await db.bookings.find_one({"chassis_number": body.chassis_number, "status": {"$ne": "Cancelled"}}, {"_id": 0})
        if dup:
            raise HTTPException(400, f"Chassis {body.chassis_number} is already booked on another lead")
        locked_chassis = body.chassis_number

    # payment_type — Token implies Booking Hold; Full → straight to Booking
    pay_type = (body.payment_type or "Token").strip()
    if pay_type not in ("Token", "Full"):
        pay_type = "Token"

    bid = str(uuid.uuid4())
    doc = {
        "id": bid,
        "lead_id": lid,
        "branch_id": lead.get("branch_id"),
        "assigned_to": lead.get("assigned_to"),
        "customer_name": lead.get("customer_name"),
        "booking_date": booking_date,
        "expected_delivery_date": body.expected_delivery_date,
        "brand_id": body.brand_id or lead.get("brand_id"),
        "model_id": body.model_id or lead.get("model_id"),
        "variant_id": body.variant_id or lead.get("variant_id"),
        "color_id": body.color_id or lead.get("color_id"),
        "final_deal_price": float(final_price),
        "booking_amount": float(body.booking_amount),
        "total_paid": 0.0,
        "pending_amount": float(final_price),
        "payment_status": "Pending",
        "status": "Pending",
        "payment_type": pay_type,
        "inventory_id": locked_inv_id,
        "chassis_number": locked_chassis,
        "finance_company": body.finance_company,
        "down_payment": body.down_payment,
        "emi": body.emi,
        "loan_status": body.loan_status,
        "exchange_final_value": body.exchange_final_value,
        "notes": body.notes,
        "receipt_file_id": None,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.bookings.insert_one(doc)
    doc.pop("_id", None)

    # Lock the inventory item
    if locked_inv_id:
        await db.inventory.update_one(
            {"id": locked_inv_id},
            {"$set": {
                "status": "booked",
                "booked_lead_id": lid,
                "booked_booking_id": bid,
                "updated_at": now_iso(),
            }},
        )

    # Auto-advance lead stage — Token → Hold; Full → Booking
    target_stage = "Hold" if pay_type == "Token" else "Booking"
    stage_order = {s: i for i, s in enumerate(STAGES)}
    if stage_order.get(lead.get("stage"), 0) < stage_order[target_stage]:
        await db.leads.update_one({"id": lid},
                                  {"$set": {"stage": target_stage, "updated_at": now_iso()}})
        await add_timeline(lid, "Stage Changed", user,
                           {"from": lead.get("stage"), "to": target_stage, "via": "booking_created"})
    await add_timeline(lid, "Booking Created", user,
                       {"booking_amount": body.booking_amount,
                        "payment_type": pay_type,
                        "chassis_number": locked_chassis,
                        "expected_delivery_date": body.expected_delivery_date})
    return doc


@api.get("/leads/{lid}/booking")
async def get_booking_for_lead(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    booking = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}},
                                         {"_id": 0})
    return booking


@api.get("/bookings")
async def list_bookings(
    branch_id: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """List bookings with their allotments. Used by reports export."""
    q: dict = {}
    if user["role"] == "admin":
        q["branch_id"] = user.get("branch_id")
    elif user["role"] in ("sales_executive", "tele_executive"):
        q["assigned_to"] = user["id"]
    if branch_id and user["role"] == "super_admin":
        q["branch_id"] = branch_id
    if status:
        q["status"] = status
    bookings = await db.bookings.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    # Attach allotment info per booking
    if bookings:
        bids = [b["id"] for b in bookings]
        allots = await db.allotments.find({"booking_id": {"$in": bids}}, {"_id": 0}).to_list(5000)
        amap = {a["booking_id"]: a for a in allots}
        for b in bookings:
            a = amap.get(b["id"])
            b["allotment"] = a
            if a:
                b["delivery_date"] = a.get("allotted_at")
                b["chassis_number"] = b.get("chassis_number") or a.get("chassis_number")
                b["engine_number"] = a.get("engine_number")
    return bookings


@api.get("/bookings/{bid}")
async def get_booking(bid: str, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return booking


@api.put("/bookings/{bid}")
async def update_booking(bid: str, body: BookingUpdate, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")

    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    # Only admin/super_admin can change booking_date or status directly
    if user["role"] == "sales_executive":
        for protected in ("booking_date", "status"):
            updates.pop(protected, None)

    # Business validations
    new_ba = updates.get("booking_amount", booking["booking_amount"])
    if new_ba > float(booking["final_deal_price"]):
        raise HTTPException(400, "Booking amount cannot exceed final deal price")
    bd = updates.get("booking_date", booking["booking_date"])
    dd = updates.get("expected_delivery_date", booking["expected_delivery_date"])
    if dd < bd:
        raise HTTPException(400, "Expected delivery date must be on or after booking date")
    if "status" in updates and updates["status"] not in BOOKING_STATUSES:
        raise HTTPException(400, "Invalid booking status")
    if "loan_status" in updates and updates["loan_status"] not in LOAN_STATUSES:
        raise HTTPException(400, "Invalid loan status")

    if updates:
        updates["updated_at"] = now_iso()
        await db.bookings.update_one({"id": bid}, {"$set": updates})
    fresh = await db.bookings.find_one({"id": bid}, {"_id": 0})
    return fresh


@api.post("/bookings/{bid}/confirm")
async def confirm_booking(bid: str, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if booking.get("status") == "Cancelled":
        raise HTTPException(400, "Booking was cancelled")
    paid = await _booking_payment_totals(bid)
    if paid < float(booking["booking_amount"]):
        raise HTTPException(400, f"Minimum booking amount ₹{booking['booking_amount']} not collected yet (paid ₹{paid})")
    await db.bookings.update_one({"id": bid}, {"$set": {"status": "Confirmed",
                                                        "confirmed_at": now_iso(),
                                                        "confirmed_by": user["id"],
                                                        "updated_at": now_iso()}})
    await add_timeline(booking["lead_id"], "Booking Confirmed", user,
                       {"total_paid": paid, "booking_amount": booking["booking_amount"]})
    return await db.bookings.find_one({"id": bid}, {"_id": 0})


@api.post("/bookings/{bid}/cancel")
async def cancel_booking(bid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if user["role"] == "admin" and (not lead or lead.get("branch_id") != user.get("branch_id")):
        raise HTTPException(403, "Cannot cancel booking outside your branch")
    await db.bookings.update_one({"id": bid}, {"$set": {"status": "Cancelled",
                                                        "cancelled_at": now_iso(),
                                                        "cancelled_by": user["id"],
                                                        "updated_at": now_iso()}})
    await add_timeline(booking["lead_id"], "Booking Cancelled", user, {})
    return await db.bookings.find_one({"id": bid}, {"_id": 0})


# ------------------ Payments ------------------

@api.post("/bookings/{bid}/payments")
async def add_payment(bid: str, body: PaymentIn, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if booking.get("status") == "Cancelled":
        raise HTTPException(400, "Booking was cancelled")
    if body.amount <= 0:
        raise HTTPException(400, "Amount must be positive")
    if body.mode not in PAYMENT_MODES:
        raise HTTPException(400, "Invalid payment mode")
    if body.payment_type and body.payment_type not in PAYMENT_TYPES:
        raise HTTPException(400, "Invalid payment type")

    # Check does not exceed final
    current_paid = await _booking_payment_totals(bid)
    if current_paid + float(body.amount) > float(booking["final_deal_price"]) + 0.01:
        raise HTTPException(400, "Total paid would exceed final deal price")

    pid = str(uuid.uuid4())
    doc = {
        "id": pid,
        "booking_id": bid,
        "lead_id": booking["lead_id"],
        "amount": float(body.amount),
        "date": body.date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "mode": body.mode,
        "payment_type": body.payment_type or "Booking",
        "notes": body.notes,
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
    }
    await db.payments.insert_one(doc)
    doc.pop("_id", None)
    fresh = await db.bookings.find_one({"id": bid}, {"_id": 0})
    fresh = await _recompute_booking(fresh)
    await add_timeline(booking["lead_id"], "Payment Added", user,
                       {"amount": body.amount, "mode": body.mode,
                        "total_paid": fresh["total_paid"]})
    return {"payment": doc, "booking": fresh}


@api.get("/bookings/{bid}/payments")
async def list_payments(bid: str, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.payments.find({"booking_id": bid}, {"_id": 0}).sort("date", -1).to_list(500)
    return items


@api.get("/bookings/{bid}/payment-summary")
async def payment_summary(bid: str, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    by_type = {k: 0.0 for k in PAYMENT_TYPES}
    async for p in db.payments.find({"booking_id": bid}, {"_id": 0, "amount": 1, "payment_type": 1}):
        t = p.get("payment_type") or "Booking"
        by_type[t] = by_type.get(t, 0.0) + float(p.get("amount") or 0)

    total_paid = sum(by_type.values())
    final_price = float(booking.get("final_deal_price") or 0)
    exchange_value = float(booking.get("exchange_adjustment") or 0)
    net_payable = max(0.0, final_price - exchange_value)
    pending = max(0.0, net_payable - total_paid)

    # Margin alert
    margin_alert = False
    days_to_delivery = None
    delivery = await db.deliveries.find_one({"lead_id": lead["id"], "status": {"$ne": "Cancelled"}}, {"_id": 0})
    if delivery and delivery.get("delivery_date"):
        try:
            ddate = datetime.strptime(delivery["delivery_date"], "%Y-%m-%d").date()
            today_d = datetime.now(timezone.utc).date()
            days_to_delivery = (ddate - today_d).days
            if 0 <= days_to_delivery <= MARGIN_ALERT_DAYS and by_type.get("Margin", 0) == 0 and pending > 0:
                margin_alert = True
        except Exception:
            pass

    return {
        "final_deal_price": final_price,
        "exchange_adjustment": exchange_value,
        "net_payable": net_payable,
        "total_paid": total_paid,
        "pending_amount": round(pending, 2),
        "by_type": by_type,
        "payment_status": booking.get("payment_status"),
        "margin_alert": margin_alert,
        "days_to_delivery": days_to_delivery,
    }


@api.get("/payments/{pid}/receipt", response_class=FastResponse)
async def payment_receipt(pid: str, request: Request,
                          authorization: Optional[str] = Header(None),
                          auth: Optional[str] = Query(None)):
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
    elif auth:
        token = auth
    else:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except Exception:
        raise HTTPException(401, "Invalid token")

    payment = await db.payments.find_one({"id": pid}, {"_id": 0})
    if not payment:
        raise HTTPException(404, "Payment not found")
    booking = await db.bookings.find_one({"id": payment["booking_id"]}, {"_id": 0}) if payment.get("booking_id") else None
    lead = await db.leads.find_one({"id": payment["lead_id"]}, {"_id": 0}) if payment.get("lead_id") else None
    branch = await db.branches.find_one({"id": (lead or {}).get("branch_id")}, {"_id": 0}) if lead else None
    brand = await db.brands.find_one({"id": (lead or {}).get("brand_id")}, {"_id": 0}) if lead else None
    model = await db.vehicle_models.find_one({"id": (lead or {}).get("model_id")}, {"_id": 0}) if lead else None

    html = f"""<!doctype html><html><head><meta charset="utf-8"><title>Payment Receipt</title>
<style>
  body {{ font-family: 'IBM Plex Sans', Arial, sans-serif; color: #09090b; padding: 40px; max-width: 640px; margin: auto; }}
  .overline {{ font-size: 10px; letter-spacing: 0.2em; color: #52525b; font-weight: 700; text-transform: uppercase; }}
  h1 {{ font-size: 22px; margin: 2px 0 0; letter-spacing: -0.02em; }}
  .row {{ display: flex; justify-content: space-between; border-bottom: 1px dashed #e4e4e7; padding: 8px 0; font-size: 14px; }}
  .total {{ background: #09090b; color: #fff; padding: 14px; margin-top: 14px; font-family: 'JetBrains Mono', monospace; font-size: 20px; text-align: center; }}
  .muted {{ color: #52525b; font-size: 12px; }}
  @media print {{ .noprint {{ display: none }} body {{ padding: 0 }} }}
</style></head><body>
<div class="noprint" style="text-align:right;margin-bottom:16px"><button onclick="window.print()">Print</button></div>
<div style="border-bottom:2px solid #09090b;padding-bottom:10px;margin-bottom:16px">
  <div class="overline">Torque Dealership CRM</div>
  <h1>Payment Receipt</h1>
  <div class="muted">#{pid[:8].upper()} · Branch: {(branch or {}).get('name','—')}</div>
</div>
<div class="row"><span class="muted">Customer</span><b>{(lead or {}).get('customer_name','—')}</b></div>
<div class="row"><span class="muted">Phone</span><span>{(lead or {}).get('phone','—')}</span></div>
<div class="row"><span class="muted">Vehicle</span><span>{(brand or {}).get('name','')} {(model or {}).get('name','')}</span></div>
<div class="row"><span class="muted">Date</span><span>{payment['date']}</span></div>
<div class="row"><span class="muted">Payment Type</span><b>{payment.get('payment_type','Booking')}</b></div>
<div class="row"><span class="muted">Mode</span><span>{payment['mode']}</span></div>
<div class="row"><span class="muted">Received By</span><span>{payment.get('created_by_name','—')}</span></div>
{'<div class="row"><span class="muted">Notes</span><span>' + (payment.get('notes') or '') + '</span></div>' if payment.get('notes') else ''}
<div class="total">₹ {payment['amount']:,.2f}</div>
<div style="margin-top:16px" class="muted">Total paid so far: ₹{(booking or {}).get('total_paid', 0):,.2f} · Pending: ₹{(booking or {}).get('pending_amount', 0):,.2f}</div>
<div style="margin-top:40px;text-align:center;color:#a1a1aa;font-size:11px">Computer-generated receipt · No signature required</div>
</body></html>"""
    return FastResponse(content=html, media_type="text/html")


# ============================================================
# Module 9 — Finance Case
# ============================================================

@api.post("/leads/{lid}/finance-case")
async def create_finance_case(lid: str, body: FinanceCaseIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    existing = await db.finance_cases.find_one({"lead_id": lid}, {"_id": 0})
    if existing:
        raise HTTPException(400, "Finance case already exists; use PUT to update")
    final_price = float(((lead.get("deal") or {}).get("final_deal_price")) or 0)
    dp = float(body.downpayment_amount or 0)
    loan_amount = max(0.0, final_price - dp) if final_price else None

    fid = str(uuid.uuid4())
    doc = {
        "id": fid, "lead_id": lid,
        "branch_id": lead.get("branch_id"),
        "finance_company": body.finance_company,
        "downpayment_amount": dp,
        "loan_amount": loan_amount,
        "emi": body.emi, "tenure": body.tenure,
        "assigned_to": body.assigned_to,
        "status": "Applied",
        "applied_at": now_iso(),
        "approved_at": None, "rejection_reason": None,
        "eligibility_notes": None,
        "downpayment_received": False,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.finance_cases.insert_one(doc)
    doc.pop("_id", None)
    await add_timeline(lid, "Finance Case Created", user,
                       {"finance_company": body.finance_company, "loan_amount": loan_amount})
    return doc


@api.get("/leads/{lid}/finance-case")
async def get_finance_case(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return await db.finance_cases.find_one({"lead_id": lid}, {"_id": 0})


@api.put("/finance-cases/{fid}")
async def update_finance_case(fid: str, body: FinanceCaseUpdate, user: dict = Depends(get_current_user)):
    fc = await db.finance_cases.find_one({"id": fid}, {"_id": 0})
    if not fc:
        raise HTTPException(404, "Finance case not found")
    lead = await db.leads.find_one({"id": fc["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}

    if "status" in updates:
        if updates["status"] not in FINANCE_STATUSES:
            raise HTTPException(400, "Invalid finance status")
        # Only admin/super_admin can approve or reject
        if updates["status"] in ("Approved", "Rejected") and user["role"] not in ("admin", "super_admin"):
            raise HTTPException(403, "Only admin can approve or reject finance")
        if updates["status"] == "Approved":
            updates["approved_at"] = now_iso()
            updates["approved_by"] = user["id"]
            updates["approved_by_name"] = user["name"]
        if updates["status"] == "Rejected" and not (updates.get("rejection_reason") or fc.get("rejection_reason")):
            raise HTTPException(400, "Rejection reason required")

    # Recompute loan amount if downpayment changes
    final_price = float(((lead.get("deal") or {}).get("final_deal_price")) or 0)
    if "downpayment_amount" in updates and final_price:
        updates["loan_amount"] = max(0.0, final_price - float(updates["downpayment_amount"]))

    if updates:
        updates["updated_at"] = now_iso()
        await db.finance_cases.update_one({"id": fid}, {"$set": updates})
        if "status" in updates:
            await add_timeline(fc["lead_id"], f"Finance {updates['status']}", user,
                               {"finance_company": fc.get("finance_company")})
    fresh = await db.finance_cases.find_one({"id": fid}, {"_id": 0})
    return fresh


# ============================================================
# Module 10 — Exchange Valuation History
# ============================================================

@api.post("/leads/{lid}/exchange-valuations")
async def add_exchange_valuation(lid: str, body: ExchangeValuationIn,
                                 user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if lead.get("purchase_type") != "Exchange Vehicle":
        raise HTTPException(400, "Lead is not an exchange case")
    doc = {
        "id": str(uuid.uuid4()),
        "lead_id": lid,
        "source": body.source,
        "value": float(body.value),
        "remarks": body.remarks,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(),
    }
    await db.exchange_valuations.insert_one(doc)
    doc.pop("_id", None)
    # Store broker_value on lead exchange for quick display if source is broker
    if body.source == "broker":
        exch = lead.get("exchange") or {}
        exch["broker_value"] = float(body.value)
        exch["broker_remarks"] = body.remarks
        await db.leads.update_one({"id": lid}, {"$set": {"exchange": exch, "updated_at": now_iso()}})
    await add_timeline(lid, "Exchange Valuation Added", user,
                       {"source": body.source, "value": body.value})
    return doc


@api.get("/leads/{lid}/exchange-valuations")
async def list_exchange_valuations(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return await db.exchange_valuations.find({"lead_id": lid}, {"_id": 0}).sort("created_at", -1).to_list(200)


IDENTITY_DOC_TYPES = {"aadhaar", "aadhaar_back", "pan", "bank_passbook", "other"}
EXCHANGE_DOC_TYPES = {"rc", "rc_front", "rc_back", "rc_pdf", "front_photo", "back_photo", "rc_book"}
LEGACY_PHOTO_TYPE = "photo"


@api.post("/leads/{lid}/exchange-photos")
async def upload_exchange_photo(lid: str, file: UploadFile = File(...),
                                doc_type: str = "photo",
                                user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    dt = (doc_type or "photo").lower()
    all_allowed = IDENTITY_DOC_TYPES | EXCHANGE_DOC_TYPES | {LEGACY_PHOTO_TYPE}
    if dt not in all_allowed:
        dt = LEGACY_PHOTO_TYPE
    ext = (file.filename or "bin").split(".")[-1].lower() if "." in (file.filename or "") else "bin"
    path = f"{APP_NAME}/leads/{lid}/docs/{dt}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    result = put_object(path, data, file.content_type or "application/octet-stream")
    file_id = str(uuid.uuid4())
    frec = {
        "id": file_id, "lead_id": lid,
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size"),
        "doc_type": f"Lead {dt.replace('_', ' ').title()}",
        "side": "identity" if dt in IDENTITY_DOC_TYPES else "exchange",
        "bucket": dt,
        "uploaded_by": user["id"], "uploaded_by_name": user["name"],
        "is_deleted": False, "created_at": now_iso(),
    }
    await db.files.insert_one(frec)

    updates: Dict[str, Any] = {"updated_at": now_iso()}
    if dt in IDENTITY_DOC_TYPES:
        # Store at top-level lead.identity_docs.<bucket>[]
        identity = lead.get("identity_docs") or {}
        bucket = identity.get(dt) or []
        bucket.append(file_id)
        identity[dt] = bucket
        updates["identity_docs"] = identity
    elif dt == LEGACY_PHOTO_TYPE:
        exch = lead.get("exchange") or {}
        photos = exch.get("photos") or []
        photos.append(file_id)
        exch["photos"] = photos
        updates["exchange"] = exch
    else:
        # Exchange doc: store at lead.exchange.documents.<bucket>[]
        exch = lead.get("exchange") or {}
        documents = exch.get("documents") or {}
        bucket_list = documents.get(dt) or []
        bucket_list.append(file_id)
        documents[dt] = bucket_list
        exch["documents"] = documents
        updates["exchange"] = exch

    await db.leads.update_one({"id": lid}, {"$set": updates})
    frec.pop("_id", None)
    refreshed = await db.leads.find_one({"id": lid}, {"_id": 0, "identity_docs": 1, "exchange": 1})
    return {
        "file_id": file_id,
        "doc_type": dt,
        "identity_docs": (refreshed or {}).get("identity_docs") or {},
        "exchange": (refreshed or {}).get("exchange") or {},
    }


@api.delete("/leads/{lid}/exchange-photos/{file_id}")
async def delete_exchange_photo(lid: str, file_id: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    changed = False
    updates: Dict[str, Any] = {}

    # identity_docs
    identity = lead.get("identity_docs") or {}
    for k, arr in list(identity.items()):
        if file_id in (arr or []):
            identity[k] = [x for x in arr if x != file_id]
            updates["identity_docs"] = identity
            changed = True

    # exchange legacy photos + documents
    exch = lead.get("exchange") or {}
    photos = exch.get("photos") or []
    if file_id in photos:
        exch["photos"] = [x for x in photos if x != file_id]
        updates["exchange"] = exch
        changed = True
    documents = exch.get("documents") or {}
    for k, arr in list(documents.items()):
        if file_id in (arr or []):
            documents[k] = [x for x in arr if x != file_id]
            exch["documents"] = documents
            updates["exchange"] = exch
            changed = True

    if changed:
        updates["updated_at"] = now_iso()
        await db.leads.update_one({"id": lid}, {"$set": updates})
        await db.files.update_one({"id": file_id}, {"$set": {"is_deleted": True}})
    return {"ok": changed}



# ============================================================
# Vehicle Allotment (Module 6)
# ============================================================

@api.post("/bookings/{bid}/allotment")
async def create_allotment(bid: str, body: AllotmentIn, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if booking.get("status") != "Confirmed":
        raise HTTPException(400, "Booking must be Confirmed before allotment")
    existing = await db.allotments.find_one({"booking_id": bid}, {"_id": 0})
    if existing:
        raise HTTPException(400, "Allotment already exists for this booking")
    chassis = (body.chassis_number or "").strip().upper() or None
    engine = (body.engine_number or "").strip().upper() or None
    # Unique chassis check (only if provided)
    if chassis:
        dup = await db.allotments.find_one({"chassis_number": chassis}, {"_id": 0})
        if dup:
            raise HTTPException(400, f"Chassis number {chassis} already assigned")

    aid = str(uuid.uuid4())
    doc = {
        "id": aid,
        "lead_id": booking["lead_id"],
        "booking_id": bid,
        "branch_id": booking.get("branch_id"),
        "chassis_number": chassis,
        "engine_number": engine,
        "status": "Allotted",
        "allotted_by": user["id"],
        "allotted_by_name": user["name"],
        "allotted_at": now_iso(),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.allotments.insert_one(doc)
    doc.pop("_id", None)

    # Auto-advance lead to Delivery
    stage_order = {s: i for i, s in enumerate(STAGES)}
    if stage_order.get(lead.get("stage"), 0) < stage_order["Delivered"]:
        await db.leads.update_one({"id": lead["id"]},
                                  {"$set": {"stage": "Delivered", "updated_at": now_iso()}})
        await add_timeline(lead["id"], "Stage Changed", user,
                           {"from": lead.get("stage"), "to": "Delivered", "via": "allotment"})
    await add_timeline(lead["id"], "Vehicle Allotted", user,
                       {"chassis_number": chassis, "engine_number": doc["engine_number"]})
    return doc


@api.get("/bookings/{bid}/allotment")
async def get_allotment(bid: str, user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": bid}, {"_id": 0})
    if not booking:
        raise HTTPException(404, "Booking not found")
    lead = await db.leads.find_one({"id": booking["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return await db.allotments.find_one({"booking_id": bid}, {"_id": 0})


@api.put("/allotments/{aid}")
async def update_allotment(aid: str, body: AllotmentUpdate,
                           user: dict = Depends(require_roles("super_admin", "admin"))):
    allot = await db.allotments.find_one({"id": aid}, {"_id": 0})
    if not allot:
        raise HTTPException(404, "Allotment not found")
    if user["role"] == "admin" and allot.get("branch_id") != user.get("branch_id"):
        raise HTTPException(403, "Cannot edit allotment outside your branch")
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "chassis_number" in updates:
        chassis = (updates["chassis_number"] or "").strip().upper() or None
        if chassis and chassis != allot.get("chassis_number"):
            dup = await db.allotments.find_one({"chassis_number": chassis}, {"_id": 0})
            if dup:
                raise HTTPException(400, f"Chassis number {chassis} already assigned")
        updates["chassis_number"] = chassis
    if "engine_number" in updates:
        updates["engine_number"] = (updates["engine_number"] or "").strip().upper() or None
    if "status" in updates and updates["status"] not in ALLOTMENT_STATUSES:
        raise HTTPException(400, "Invalid status")
    if updates:
        updates["updated_at"] = now_iso()
        await db.allotments.update_one({"id": aid}, {"$set": updates})
    return await db.allotments.find_one({"id": aid}, {"_id": 0})


# ============================================================


# ============================================================
# Module 7 — Documents (with Gemini OCR)
# ============================================================

def _mask_doc_number(doc_type: Optional[str], n: Optional[str]) -> Optional[str]:
    if not n:
        return n
    s = str(n).replace(" ", "")
    if doc_type == "Aadhaar Card" and len(s) >= 4:
        return "XXXX XXXX " + s[-4:]
    if doc_type == "PAN Card" and len(s) >= 4:
        return s[:2] + "XXXX" + s[-2:]
    return n


def _present_document(d: dict, mask: bool = True) -> dict:
    out = dict(d)
    out.pop("_id", None)
    if mask and out.get("doc_number"):
        out["doc_number_masked"] = _mask_doc_number(out.get("doc_type"), out["doc_number"])
        if out.get("doc_type") in ("Aadhaar Card", "PAN Card"):
            out["doc_number"] = out["doc_number_masked"]
    return out


async def _upload_doc_file(lid: str, doc_type: str, upload: UploadFile, user: dict, side: str) -> Optional[dict]:
    if not upload:
        return None
    ext = (upload.filename or "bin").split(".")[-1].lower() if "." in (upload.filename or "") else "bin"
    path = f"{APP_NAME}/leads/{lid}/documents/{uuid.uuid4()}.{ext}"
    data = await upload.read()
    result = put_object(path, data, upload.content_type or "application/octet-stream")
    file_id = str(uuid.uuid4())
    frec = {
        "id": file_id, "lead_id": lid,
        "storage_path": result["path"],
        "original_filename": upload.filename,
        "content_type": upload.content_type,
        "size": result.get("size"),
        "doc_type": doc_type, "side": side,
        "uploaded_by": user["id"], "uploaded_by_name": user["name"],
        "is_deleted": False, "created_at": now_iso(),
    }
    await db.files.insert_one(frec)
    frec.pop("_id", None)
    return frec


@api.post("/leads/{lid}/documents")
async def upload_document(
    lid: str,
    doc_type: str = Form(...),
    doc_number: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    front: UploadFile = File(...),
    back: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user),
):
    if doc_type not in DOC_TYPES:
        raise HTTPException(400, "Invalid doc_type")
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")

    # Supersede existing latest of same doc_type
    await db.documents.update_many(
        {"lead_id": lid, "doc_type": doc_type, "is_latest": True},
        {"$set": {"is_latest": False}},
    )
    prev = await db.documents.count_documents({"lead_id": lid, "doc_type": doc_type})

    front_file = await _upload_doc_file(lid, doc_type, front, user, "front")
    back_file = await _upload_doc_file(lid, doc_type, back, user, "back") if back else None

    did = str(uuid.uuid4())
    doc = {
        "id": did, "lead_id": lid,
        "branch_id": lead.get("branch_id"),
        "customer_name": lead.get("customer_name"),
        "doc_type": doc_type,
        "doc_number": (doc_number or "").strip() or None,
        "front_file_id": front_file["id"] if front_file else None,
        "back_file_id": back_file["id"] if back_file else None,
        "extracted": {}, "ocr_ran": False, "ocr_at": None,
        "status": "Pending",
        "version": prev + 1, "is_latest": True,
        "verified_by": None, "verified_by_name": None, "verified_at": None,
        "rejection_reason": None,
        "notes": notes,
        "uploaded_by": user["id"], "uploaded_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.documents.insert_one(doc)
    doc.pop("_id", None)
    await add_timeline(lid, "Document Uploaded", user,
                       {"doc_type": doc_type, "version": doc["version"]})
    return _present_document(doc)


@api.get("/leads/{lid}/documents")
async def list_documents(lid: str, include_history: bool = False,
                         user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    q: Dict[str, Any] = {"lead_id": lid}
    if not include_history:
        q["is_latest"] = True
    items = await db.documents.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [_present_document(d) for d in items]


@api.put("/documents/{did}")
async def update_document(did: str, body: DocumentUpdate, user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    lead = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    updates: Dict[str, Any] = {}
    if body.doc_type and body.doc_type in DOC_TYPES:
        updates["doc_type"] = body.doc_type
    if body.doc_number is not None:
        updates["doc_number"] = body.doc_number.strip() or None
    if body.extracted is not None:
        updates["extracted"] = body.extracted.model_dump(exclude_none=False)
    if body.notes is not None:
        updates["notes"] = body.notes
    if updates:
        updates["updated_at"] = now_iso()
        await db.documents.update_one({"id": did}, {"$set": updates})
    fresh = await db.documents.find_one({"id": did}, {"_id": 0})
    return _present_document(fresh)


@api.post("/documents/{did}/verify")
async def verify_document(did: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    doc = await db.documents.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    lead = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
    if user["role"] == "admin" and (not lead or lead.get("branch_id") != user.get("branch_id")):
        raise HTTPException(403, "Out of branch")
    await db.documents.update_one({"id": did}, {"$set": {
        "status": "Verified",
        "verified_by": user["id"], "verified_by_name": user["name"],
        "verified_at": now_iso(), "rejection_reason": None,
        "updated_at": now_iso(),
    }})
    await add_timeline(doc["lead_id"], "Document Verified", user, {"doc_type": doc["doc_type"]})
    fresh = await db.documents.find_one({"id": did}, {"_id": 0})
    return _present_document(fresh)


@api.post("/documents/{did}/reject")
async def reject_document(did: str, body: DocumentReject,
                          user: dict = Depends(require_roles("super_admin", "admin"))):
    doc = await db.documents.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    lead = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
    if user["role"] == "admin" and (not lead or lead.get("branch_id") != user.get("branch_id")):
        raise HTTPException(403, "Out of branch")
    await db.documents.update_one({"id": did}, {"$set": {
        "status": "Rejected",
        "rejection_reason": body.reason,
        "verified_by": user["id"], "verified_by_name": user["name"],
        "verified_at": now_iso(),
        "updated_at": now_iso(),
    }})
    await add_timeline(doc["lead_id"], "Document Rejected", user,
                       {"doc_type": doc["doc_type"], "reason": body.reason})
    fresh = await db.documents.find_one({"id": did}, {"_id": 0})
    return _present_document(fresh)


@api.get("/documents/{did}/duplicates")
async def duplicate_documents(did: str, user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    lead = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    number = doc.get("doc_number")
    if not number:
        return []
    dups = await db.documents.find({
        "doc_type": doc["doc_type"],
        "doc_number": number,
        "lead_id": {"$ne": doc["lead_id"]},
    }, {"_id": 0}).to_list(50)
    return [_present_document(d) for d in dups]


@api.post("/documents/{did}/ocr")
async def run_ocr(did: str, user: dict = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Document not found")
    lead = await db.leads.find_one({"id": doc["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    front = await db.files.find_one({"id": doc.get("front_file_id")}, {"_id": 0}) if doc.get("front_file_id") else None
    if not front:
        raise HTTPException(400, "No front image to OCR")

    try:
        import base64 as _b64
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent

        front_bytes, front_ctype = get_object(front["storage_path"])
        b64_front = _b64.b64encode(front_bytes).decode()
        images = [ImageContent(image_base64=b64_front)]
        if doc.get("back_file_id"):
            back = await db.files.find_one({"id": doc["back_file_id"]}, {"_id": 0})
            if back:
                back_bytes, _ = get_object(back["storage_path"])
                images.append(ImageContent(image_base64=_b64.b64encode(back_bytes).decode()))

        prompt = (
            f"You are an OCR assistant for an Indian two-wheeler dealership. "
            f"Extract fields from this {doc['doc_type']} image(s). "
            "Respond with ONLY a strict JSON object (no markdown, no commentary) with keys: "
            '"document_number", "name", "address", "chassis_number", "engine_number", '
            '"vehicle_model", "variant", "confidence_score" (0.0 to 1.0). '
            "Use empty string for missing fields."
        )
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"ocr-{did}",
            system_message="You extract structured fields from document images and return strict JSON only.",
        ).with_model("gemini", "gemini-2.5-flash")
        msg = UserMessage(text=prompt, file_contents=images)
        raw = await chat.send_message(msg)
        raw = (raw or "").strip()
        if raw.startswith("```"):
            raw = raw.strip("`")
            if raw.lower().startswith("json"):
                raw = raw[4:].lstrip()
        import json as _json
        try:
            data = _json.loads(raw)
        except Exception:
            # Try to locate a JSON object
            start = raw.find("{")
            end = raw.rfind("}")
            data = _json.loads(raw[start:end + 1]) if start >= 0 and end > start else {}

        extracted = {
            "document_number": (data.get("document_number") or "").strip() or None,
            "name": (data.get("name") or "").strip() or None,
            "address": (data.get("address") or "").strip() or None,
            "chassis_number": (data.get("chassis_number") or "").strip().upper() or None,
            "engine_number": (data.get("engine_number") or "").strip().upper() or None,
            "vehicle_model": (data.get("vehicle_model") or "").strip() or None,
            "variant": (data.get("variant") or "").strip() or None,
            "confidence_score": float(data.get("confidence_score") or 0.0),
        }
        await db.documents.update_one({"id": did}, {"$set": {
            "extracted": extracted,
            "ocr_ran": True,
            "ocr_at": now_iso(),
            "updated_at": now_iso(),
        }})
        if not doc.get("doc_number") and extracted.get("document_number"):
            await db.documents.update_one({"id": did},
                                          {"$set": {"doc_number": extracted["document_number"]}})
        fresh = await db.documents.find_one({"id": did}, {"_id": 0})
        return _present_document(fresh)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("OCR failed")
        raise HTTPException(500, f"OCR failed: {e}")


def _stage_doc_requirements(stage: str, lead_docs: List[dict]) -> List[str]:
    required = DOC_REQUIREMENTS.get(stage, [])
    verified_types = {d["doc_type"] for d in lead_docs
                      if d.get("is_latest") and d.get("status") == "Verified"}
    missing = [t for t in required if t not in verified_types]
    return missing


# ============================================================
# Module 8 — Delivery
# ============================================================

async def _render_template(body: str, variables: Dict[str, Any]) -> str:
    out = body or ""
    for k, v in (variables or {}).items():
        out = out.replace("{{" + str(k) + "}}", str(v if v is not None else ""))
    return out


async def _lead_variables(lead: dict, extra: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    brand = await db.brands.find_one({"id": lead.get("brand_id")}, {"_id": 0}) if lead.get("brand_id") else None
    model = await db.vehicle_models.find_one({"id": lead.get("model_id")}, {"_id": 0}) if lead.get("model_id") else None
    branch = await db.branches.find_one({"id": lead.get("branch_id")}, {"_id": 0}) if lead.get("branch_id") else None
    exec_u = await db.users.find_one({"id": lead.get("assigned_to")}, {"_id": 0, "password_hash": 0}) if lead.get("assigned_to") else None
    deal = lead.get("deal") or {}
    vars_ = {
        "customer_name": lead.get("customer_name", ""),
        "phone": lead.get("phone", ""),
        "vehicle": " ".join(filter(None, [(brand or {}).get("name"), (model or {}).get("name")])) or "your vehicle",
        "brand": (brand or {}).get("name", ""),
        "model": (model or {}).get("name", ""),
        "branch": (branch or {}).get("name", ""),
        "sales_exec": (exec_u or {}).get("name", "our team"),
        "stage": lead.get("stage", ""),
        "deal_amount": deal.get("final_deal_price") or deal.get("offered_price") or "",
        "price": deal.get("offered_price") or "",
        "discount": deal.get("discount") or "",
    }
    vars_.update(extra or {})
    return vars_


async def _is_opted_out(lead_id: str, phone: Optional[str]) -> bool:
    q = {"$or": [{"lead_id": lead_id}]}
    if phone:
        q["$or"].append({"phone": phone})
    return bool(await db.wa_optouts.find_one(q))


def _conditions_match(lead: dict, conditions: Dict[str, Any]) -> bool:
    for k, v in (conditions or {}).items():
        if v is None or v == "" or v == []:
            continue
        lv = lead.get(k)
        if isinstance(v, list):
            if lv not in v:
                return False
        else:
            if lv != v:
                return False
    return True


async def _queue_message(
    lead: dict,
    *,
    template_id: Optional[str] = None,
    content: Optional[str] = None,
    message_type: str = "text",
    media_url: Optional[str] = None,
    trigger: str = "manual",
    variables: Optional[Dict[str, Any]] = None,
    user: Optional[dict] = None,
    campaign_id: Optional[str] = None,
    scheduled_at: Optional[str] = None,
    skip_guards: bool = False,
) -> Optional[dict]:
    if await _is_opted_out(lead["id"], lead.get("phone")):
        return None
    rendered = content
    if template_id and not rendered:
        tpl = await db.wa_templates.find_one({"id": template_id, "active": True}, {"_id": 0})
        if not tpl:
            return None
        message_type = tpl.get("message_type", message_type)
        media_url = media_url or tpl.get("media_url")
        all_vars = await _lead_variables(lead, variables)
        rendered = await _render_template(tpl.get("body", ""), all_vars)

    # Safety: duplicate + rate-limit guard (skippable for campaign bulk sends which
    # run their own 24h dedupe inside campaign_send)
    if not skip_guards:
        now_dt = datetime.now(timezone.utc)
        # 1) Duplicate: identical outbound content to same lead in last 60s
        dup_since = (now_dt - timedelta(seconds=60)).isoformat()
        dup = await db.wa_messages.find_one({
            "lead_id": lead["id"],
            "direction": "outbound",
            "content": rendered or "",
            "created_at": {"$gte": dup_since},
        })
        if dup:
            raise HTTPException(429, "Duplicate message: same content sent in the last 60s")
        # 2) Rate limit: max 10 outbound messages per lead per minute
        rate_since = (now_dt - timedelta(seconds=60)).isoformat()
        recent_count = await db.wa_messages.count_documents({
            "lead_id": lead["id"],
            "direction": "outbound",
            "created_at": {"$gte": rate_since},
        })
        if recent_count >= 10:
            raise HTTPException(429, "Rate limit exceeded: max 10 messages per minute for this lead")

    doc = {
        "id": str(uuid.uuid4()),
        "lead_id": lead["id"],
        "branch_id": lead.get("branch_id"),
        "phone": lead.get("phone"),
        "template_id": template_id,
        "message_type": message_type,
        "content": rendered or "",
        "media_url": media_url,
        "direction": "outbound",
        "status": "PENDING",
        "trigger_source": trigger,
        "campaign_id": campaign_id,
        "retry_count": 0,
        "reply_tag": None,
        "scheduled_at": scheduled_at,
        "created_by": (user or {}).get("id"),
        "created_by_name": (user or {}).get("name"),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.wa_messages.insert_one(doc)
    # In real wiring a Twilio/WATI worker would send here. For now, auto-mark SENT.
    await db.wa_messages.update_one({"id": doc["id"]},
                                    {"$set": {"status": "SENT", "sent_at": now_iso()}})
    doc["status"] = "SENT"
    doc.pop("_id", None)
    return doc


async def _log_whatsapp(lead: dict, intent: str, payload: dict, user: Optional[dict] = None):
    """Legacy helper — kept for backward compat; now also queues via engine."""
    # Keep the old log table for older UI references (still tail-read by DeliverySection)
    await db.whatsapp_logs.insert_one({
        "id": str(uuid.uuid4()),
        "lead_id": lead["id"],
        "phone": lead.get("phone"),
        "intent": intent,
        "payload": payload,
        "status": "LOGGED",
        "actor_id": (user or {}).get("id"),
        "created_at": now_iso(),
    })


async def fire_event(event: str, lead: dict, user: Optional[dict] = None,
                     extra: Optional[Dict[str, Any]] = None):
    """Look up matching rules and queue WA messages."""
    if not lead or not lead.get("phone"):
        return
    rules = await db.automation_rules.find({"event": event, "active": True}, {"_id": 0}).to_list(200)
    for r in rules:
        if not _conditions_match(lead, r.get("conditions") or {}):
            continue
        scheduled_at = None
        if r.get("delay_minutes"):
            scheduled_at = (datetime.now(timezone.utc) + timedelta(minutes=int(r["delay_minutes"]))).isoformat()
        await _queue_message(
            lead,
            template_id=r["template_id"],
            trigger=event,
            variables=extra,
            user=user,
            scheduled_at=scheduled_at,
        )


@api.get("/leads/{lid}/whatsapp-logs")
async def list_whatsapp_logs(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.whatsapp_logs.find({"lead_id": lid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return items


# ---------- WA Templates ----------

@api.get("/wa-templates")
async def list_templates(user: dict = Depends(get_current_user)):
    return await db.wa_templates.find({}, {"_id": 0}).sort("name", 1).to_list(500)


@api.post("/wa-templates")
async def create_template(body: WATemplateIn, user: dict = Depends(require_roles("super_admin", "admin"))):
    if body.message_type not in WA_MESSAGE_TYPES:
        raise HTTPException(400, "Invalid message_type")
    doc = body.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_by": user["id"],
                "created_at": now_iso(), "updated_at": now_iso()})
    await db.wa_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.get("/settings/integrations")
async def get_integrations(user: dict = Depends(require_roles("super_admin", "admin"))):
    doc = await db.settings.find_one({"_key": "integrations"}, {"_id": 0}) or {}
    # Never return the actual api key value verbatim in plaintext; show masked if set
    api_key = doc.get("elevenza_api_key") or ""
    masked = ("•" * max(0, len(api_key) - 4)) + api_key[-4:] if api_key else ""
    return {
        "elevenza_api_key_set": bool(api_key),
        "elevenza_api_key_masked": masked,
        "elevenza_sender_id": doc.get("elevenza_sender_id") or "",
        "triggers": doc.get("triggers") or {
            "inquiry_created": {"enabled": False, "template_id": None},
            "delivery_completed": {"enabled": False, "template_id": None},
            "feedback_reminder": {"enabled": False, "template_id": None},
        },
    }


class IntegrationSettings(BaseModel):
    elevenza_api_key: Optional[str] = None
    elevenza_sender_id: Optional[str] = None
    triggers: Optional[Dict[str, Any]] = None


@api.put("/settings/integrations")
async def put_integrations(body: IntegrationSettings, user: dict = Depends(require_roles("super_admin"))):
    doc = await db.settings.find_one({"_key": "integrations"}, {"_id": 0}) or {"_key": "integrations"}
    if body.elevenza_api_key is not None:
        doc["elevenza_api_key"] = body.elevenza_api_key
    if body.elevenza_sender_id is not None:
        doc["elevenza_sender_id"] = body.elevenza_sender_id
    if body.triggers is not None:
        doc["triggers"] = body.triggers
    doc["updated_at"] = now_iso()
    await db.settings.update_one({"_key": "integrations"}, {"$set": doc}, upsert=True)
    return {"ok": True}


class BulkSendIn(BaseModel):
    lead_ids: List[str]
    template_id: str


@api.post("/wa/bulk-send")
async def wa_bulk_send(body: BulkSendIn, user: dict = Depends(require_roles("super_admin", "admin"))):
    tpl = await db.wa_templates.find_one({"id": body.template_id}, {"_id": 0})
    if not tpl:
        raise HTTPException(404, "Template not found")
    sent = 0
    for lid in body.lead_ids:
        lead = await db.leads.find_one({"id": lid}, {"_id": 0})
        if not lead:
            continue
        if user["role"] == "admin" and lead.get("branch_id") != user.get("branch_id"):
            continue
        # Queue a mock outbound send log
        await db.wa_messages.insert_one({
            "id": str(uuid.uuid4()),
            "lead_id": lid,
            "template_id": body.template_id,
            "direction": "out",
            "status": "queued_mock",
            "body": (tpl.get("body") or "").replace("{{customer_name}}", lead.get("customer_name") or ""),
            "created_at": now_iso(),
            "created_by": user["id"],
        })
        sent += 1
    return {"ok": True, "sent": sent}


@api.put("/wa-templates/{tid}")
async def update_template(tid: str, body: WATemplateUpdate,
                          user: dict = Depends(require_roles("super_admin", "admin"))):
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "message_type" in updates and updates["message_type"] not in WA_MESSAGE_TYPES:
        raise HTTPException(400, "Invalid message_type")
    if updates:
        updates["updated_at"] = now_iso()
        await db.wa_templates.update_one({"id": tid}, {"$set": updates})
    return await db.wa_templates.find_one({"id": tid}, {"_id": 0})


@api.delete("/wa-templates/{tid}")
async def delete_template(tid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.wa_templates.delete_one({"id": tid})
    return {"ok": True}


# ---------- Automation Rules ----------

@api.get("/automation-rules")
async def list_rules(user: dict = Depends(get_current_user)):
    return await db.automation_rules.find({}, {"_id": 0}).sort("event", 1).to_list(500)


@api.post("/automation-rules")
async def create_rule(body: AutomationRuleIn,
                      user: dict = Depends(require_roles("super_admin", "admin"))):
    if body.event not in WA_EVENTS:
        raise HTTPException(400, "Invalid event")
    doc = body.model_dump()
    doc.update({"id": str(uuid.uuid4()), "created_at": now_iso(), "updated_at": now_iso()})
    await db.automation_rules.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/automation-rules/{rid}")
async def update_rule(rid: str, body: AutomationRuleUpdate,
                      user: dict = Depends(require_roles("super_admin", "admin"))):
    updates = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "event" in updates and updates["event"] not in WA_EVENTS:
        raise HTTPException(400, "Invalid event")
    if updates:
        updates["updated_at"] = now_iso()
        await db.automation_rules.update_one({"id": rid}, {"$set": updates})
    return await db.automation_rules.find_one({"id": rid}, {"_id": 0})


@api.delete("/automation-rules/{rid}")
async def delete_rule(rid: str, user: dict = Depends(require_roles("super_admin"))):
    await db.automation_rules.delete_one({"id": rid})
    return {"ok": True}


# ---------- Lead-level messaging ----------

@api.get("/leads/{lid}/wa-messages")
async def list_lead_messages(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    items = await db.wa_messages.find({"lead_id": lid}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return items


@api.post("/leads/{lid}/wa-messages")
async def send_manual(lid: str, body: ManualMessageIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if not body.template_id and not body.content:
        raise HTTPException(400, "Provide template_id or content")
    if body.message_type not in WA_MESSAGE_TYPES:
        raise HTTPException(400, "Invalid message_type")
    # Pre-checks to disambiguate failure modes before _queue_message returns None
    if await _is_opted_out(lead["id"], lead.get("phone")):
        raise HTTPException(403, "Lead has opted out of WhatsApp communication")
    if body.template_id:
        tpl_exists = await db.wa_templates.find_one(
            {"id": body.template_id, "active": True}, {"_id": 0, "id": 1}
        )
        if not tpl_exists:
            raise HTTPException(400, "Invalid or inactive template_id")
    doc = await _queue_message(
        lead,
        template_id=body.template_id,
        content=body.content,
        message_type=body.message_type,
        media_url=body.media_url,
        variables=body.variables,
        trigger="manual",
        user=user,
    )
    if not doc:
        raise HTTPException(400, "Message not sent")
    return doc


@api.post("/leads/{lid}/wa-inbound")
async def receive_inbound(lid: str, body: InboundMessageIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if body.reply_tag and body.reply_tag not in WA_REPLY_TAGS:
        raise HTTPException(400, "Invalid reply_tag")
    doc = {
        "id": str(uuid.uuid4()),
        "lead_id": lid,
        "branch_id": lead.get("branch_id"),
        "phone": lead.get("phone"),
        "template_id": None,
        "message_type": "text",
        "content": body.content,
        "media_url": None,
        "direction": "inbound",
        "status": "READ",
        "trigger_source": "inbound",
        "campaign_id": None,
        "retry_count": 0,
        "reply_tag": body.reply_tag,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.wa_messages.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.post("/wa-messages/{mid}/mark")
async def mark_message(mid: str, body: MessageMarkIn, user: dict = Depends(get_current_user)):
    if body.status not in WA_STATUSES:
        raise HTTPException(400, "Invalid status")
    msg = await db.wa_messages.find_one({"id": mid}, {"_id": 0})
    if not msg:
        raise HTTPException(404, "Message not found")
    lead = await db.leads.find_one({"id": msg["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    updates = {"status": body.status, "updated_at": now_iso()}
    if body.status == "SENT":
        updates["sent_at"] = now_iso()
    if body.status == "DELIVERED":
        updates["delivered_at"] = now_iso()
    if body.status == "READ":
        updates["read_at"] = now_iso()
    if body.status == "FAILED":
        updates["failed_at"] = now_iso()
    await db.wa_messages.update_one({"id": mid}, {"$set": updates})
    return await db.wa_messages.find_one({"id": mid}, {"_id": 0})


@api.post("/wa-messages/{mid}/retry")
async def retry_message(mid: str, user: dict = Depends(get_current_user)):
    msg = await db.wa_messages.find_one({"id": mid}, {"_id": 0})
    if not msg:
        raise HTTPException(404, "Message not found")
    lead = await db.leads.find_one({"id": msg["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if msg.get("retry_count", 0) >= WA_MAX_RETRIES:
        raise HTTPException(400, "Max retries reached")
    await db.wa_messages.update_one({"id": mid}, {
        "$inc": {"retry_count": 1},
        "$set": {"status": "SENT", "updated_at": now_iso(), "sent_at": now_iso()},
    })
    return await db.wa_messages.find_one({"id": mid}, {"_id": 0})


@api.post("/leads/{lid}/wa-optout")
async def optout(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    await db.wa_optouts.update_one(
        {"lead_id": lid},
        {"$set": {"id": str(uuid.uuid4()), "lead_id": lid, "phone": lead.get("phone"),
                  "created_at": now_iso(), "created_by": user["id"]}},
        upsert=True,
    )
    return {"opted_out": True}


@api.delete("/leads/{lid}/wa-optout")
async def optin(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    await db.wa_optouts.delete_many({"lead_id": lid})
    return {"opted_out": False}


@api.get("/leads/{lid}/wa-optout")
async def optout_status(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    exists = await db.wa_optouts.find_one({"lead_id": lid}, {"_id": 0})
    return {"opted_out": bool(exists)}


# ============================================================
# Module 12 — Campaigns
# ============================================================

def _campaign_filter_to_query(target: dict, scope: Dict[str, Any]) -> Dict[str, Any]:
    q: Dict[str, Any] = {**scope}
    t = target or {}
    if t.get("stages"):
        audience = t.get("audience") or "leads"
        if audience == "past_buyers":
            q["stage"] = {"$in": ["Delivered", "RTO", "Delivered"]}
        elif audience == "all":
            pass
        else:
            q["stage"] = {"$in": t["stages"]}
    if t.get("priorities"):
        q["priority"] = {"$in": t["priorities"]}
    if t.get("sources"):
        q["source"] = {"$in": t["sources"]}
    if t.get("branch_ids"):
        q["branch_id"] = {"$in": t["branch_ids"]}
    if t.get("purchase_types"):
        q["purchase_type"] = {"$in": t["purchase_types"]}
    return q


async def _campaign_audience_query(camp: dict, user: dict) -> Dict[str, Any]:
    scope: Dict[str, Any] = {}
    if user["role"] == "admin":
        scope["branch_id"] = user.get("branch_id")
    return _campaign_filter_to_query(camp.get("target") or {}, scope)


@api.get("/campaigns")
async def list_campaigns(user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = {}
    if user["role"] == "admin":
        # Admin sees: campaigns they created, campaigns scoped to their branch,
        # and global (branch_scope=None) campaigns created by super_admin
        q["$or"] = [
            {"created_by": user["id"]},
            {"branch_scope": user.get("branch_id")},
            {"branch_scope": None},
        ]
    elif user["role"] == "sales_executive":
        raise HTTPException(403, "Access denied")
    items = await db.campaigns.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.post("/campaigns")
async def create_campaign(body: CampaignIn,
                          user: dict = Depends(require_roles("super_admin", "admin"))):
    if body.campaign_type not in CAMPAIGN_TYPES:
        raise HTTPException(400, "Invalid campaign_type")
    if body.message_type not in WA_MESSAGE_TYPES:
        raise HTTPException(400, "Invalid message_type")
    if not body.template_id and not body.content:
        raise HTTPException(400, "Provide template_id or content")
    doc = {
        "id": str(uuid.uuid4()),
        "name": body.name,
        "campaign_type": body.campaign_type,
        "template_id": body.template_id,
        "message_type": body.message_type,
        "content": body.content,
        "media_url": body.media_url,
        "scheduled_at": body.scheduled_at,
        "target": body.target.model_dump() if body.target else {},
        "status": "Scheduled" if body.scheduled_at else "Draft",
        "branch_scope": user.get("branch_id") if user["role"] == "admin" else None,
        "stats": {"sent": 0, "delivered": 0, "read": 0, "failed": 0,
                  "responses": 0, "conversions": 0, "queued": 0},
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.campaigns.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/campaigns/{cid}")
async def update_campaign(cid: str, body: CampaignUpdate,
                          user: dict = Depends(require_roles("super_admin", "admin"))):
    updates: Dict[str, Any] = {k: v for k, v in body.model_dump(exclude_unset=True).items() if v is not None}
    if "target" in updates and isinstance(updates["target"], dict):
        pass
    if "status" in updates and updates["status"] not in CAMPAIGN_STATUSES:
        raise HTTPException(400, "Invalid status")
    if updates:
        updates["updated_at"] = now_iso()
        await db.campaigns.update_one({"id": cid}, {"$set": updates})
    return await db.campaigns.find_one({"id": cid}, {"_id": 0})


@api.delete("/campaigns/{cid}")
async def delete_campaign(cid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    await db.campaigns.delete_one({"id": cid})
    return {"ok": True}


@api.post("/campaigns/{cid}/preview")
async def campaign_preview(cid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    camp = await db.campaigns.find_one({"id": cid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    q = await _campaign_audience_query(camp, user)
    count = await db.leads.count_documents(q)
    sample = await db.leads.find(q, {"_id": 0, "id": 1, "customer_name": 1, "phone": 1,
                                     "stage": 1, "source": 1}).limit(10).to_list(10)
    return {"audience_count": count, "sample": sample}


@api.post("/campaigns/{cid}/send")
async def campaign_send(cid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    camp = await db.campaigns.find_one({"id": cid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    if camp.get("status") in ("Running", "Completed"):
        raise HTTPException(400, f"Campaign already {camp['status']}")
    q = await _campaign_audience_query(camp, user)
    await db.campaigns.update_one({"id": cid},
                                  {"$set": {"status": "Running", "started_at": now_iso()}})

    sent = 0
    async for lead in db.leads.find(q, {"_id": 0}):
        if not lead.get("phone"):
            continue
        # Duplicate protection: skip if same campaign messaged this lead in last 24h
        dup = await db.wa_messages.find_one({"lead_id": lead["id"], "campaign_id": cid})
        if dup:
            continue
        await _queue_message(
            lead,
            template_id=camp.get("template_id"),
            content=camp.get("content"),
            message_type=camp.get("message_type", "text"),
            media_url=camp.get("media_url"),
            trigger="campaign",
            user=user,
            campaign_id=cid,
            skip_guards=True,
        )
        sent += 1

    await db.campaigns.update_one({"id": cid}, {"$set": {
        "status": "Completed",
        "completed_at": now_iso(),
        "stats.queued": sent,
    }})
    return {"queued": sent}


@api.get("/campaigns/{cid}/stats")
async def campaign_stats(cid: str, user: dict = Depends(require_roles("super_admin", "admin"))):
    camp = await db.campaigns.find_one({"id": cid}, {"_id": 0})
    if not camp:
        raise HTTPException(404, "Campaign not found")
    stats = {"queued": 0, "sent": 0, "delivered": 0, "read": 0, "failed": 0}
    pipeline = [{"$match": {"campaign_id": cid}},
                {"$group": {"_id": "$status", "n": {"$sum": 1}}}]
    async for row in db.wa_messages.aggregate(pipeline):
        s = row["_id"]
        stats["queued"] += row["n"]
        if s == "SENT": stats["sent"] = row["n"]
        elif s == "DELIVERED": stats["delivered"] = row["n"]
        elif s == "READ": stats["read"] = row["n"]
        elif s == "FAILED": stats["failed"] = row["n"]

    # Responses: inbound messages by leads who got this campaign
    lead_ids = await db.wa_messages.distinct("lead_id", {"campaign_id": cid})
    responses = await db.wa_messages.count_documents({
        "lead_id": {"$in": lead_ids}, "direction": "inbound",
    })
    conversions = await db.leads.count_documents({
        "id": {"$in": lead_ids},
        "stage": {"$in": ["Booking", "Delivered", "RTO", "Delivered"]},
    })
    stats["responses"] = responses
    stats["conversions"] = conversions
    return {**camp.get("stats", {}), **stats}


@api.post("/leads/{lid}/delivery")
async def create_delivery(lid: str, body: DeliveryIn, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    existing = await db.deliveries.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}},
                                            {"_id": 0})
    if existing:
        raise HTTPException(400, "Delivery already scheduled")

    booking = await db.bookings.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}},
                                         {"_id": 0})
    allotment = await db.allotments.find_one({"booking_id": booking["id"]}, {"_id": 0}) if booking else None

    # Instant-delivery bypass requires admin/super_admin + reason
    if body.instant_bypass:
        if user["role"] not in ("super_admin", "admin"):
            raise HTTPException(403, "Instant delivery bypass requires admin approval")
        if not body.bypass_reason:
            raise HTTPException(400, "Bypass reason required")
    else:
        if not booking or booking["status"] != "Confirmed":
            raise HTTPException(400, "Booking must be Confirmed before delivery")
        if not allotment:
            raise HTTPException(400, "Vehicle must be allotted before delivery")

    did = str(uuid.uuid4())
    doc = {
        "id": did, "lead_id": lid,
        "booking_id": booking["id"] if booking else None,
        "allotment_id": allotment["id"] if allotment else None,
        "branch_id": lead.get("branch_id"),
        "delivery_date": body.delivery_date,
        "time_slot": body.time_slot,
        "delivered_by": body.delivered_by or user["id"],
        "status": "Scheduled",
        "checklist": {"payment_completed": False, "documents_verified": False,
                      "vehicle_ready": False, "accessories_ready": False},
        "accessories": [],
        "otp_hash": None, "otp_expires": None, "otp_verified": False, "otp_verified_at": None,
        "instant_bypass": bool(body.instant_bypass),
        "bypass_reason": body.bypass_reason,
        "bypass_approved_by": user["id"] if body.instant_bypass else None,
        "notes": body.notes,
        "customer_name": lead.get("customer_name"),
        "customer_phone": lead.get("phone"),
        "chassis_number": (allotment or {}).get("chassis_number"),
        "engine_number": (allotment or {}).get("engine_number"),
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.deliveries.insert_one(doc)
    doc.pop("_id", None)
    await add_timeline(lid, "Delivery Scheduled", user,
                       {"delivery_date": body.delivery_date, "instant_bypass": bool(body.instant_bypass)})
    await _log_whatsapp(lead, "delivery_scheduled",
                        {"delivery_date": body.delivery_date, "time_slot": body.time_slot}, user)
    return doc


@api.get("/leads/{lid}/delivery")
async def get_delivery_for_lead(lid: str, user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lid}, {"_id": 0})
    if not lead:
        raise HTTPException(404, "Lead not found")
    if not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    return await db.deliveries.find_one({"lead_id": lid, "status": {"$ne": "Cancelled"}},
                                        {"_id": 0})


@api.put("/deliveries/{did}")
async def update_delivery(did: str, body: DeliveryUpdate, user: dict = Depends(get_current_user)):
    delivery = await db.deliveries.find_one({"id": did}, {"_id": 0})
    if not delivery:
        raise HTTPException(404, "Delivery not found")
    lead = await db.leads.find_one({"id": delivery["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    updates = body.model_dump(exclude_unset=True)
    if "status" in updates and updates["status"] not in DELIVERY_STATUSES:
        raise HTTPException(400, "Invalid status")
    if "accessories" in updates and updates["accessories"] is not None:
        updates["accessories"] = [a for a in updates["accessories"]]
    if updates:
        updates["updated_at"] = now_iso()
        await db.deliveries.update_one({"id": did}, {"$set": updates})
    fresh = await db.deliveries.find_one({"id": did}, {"_id": 0})
    return fresh


@api.post("/deliveries/{did}/otp-generate")
async def otp_generate(did: str, user: dict = Depends(get_current_user)):
    delivery = await db.deliveries.find_one({"id": did}, {"_id": 0})
    if not delivery:
        raise HTTPException(404, "Delivery not found")
    lead = await db.leads.find_one({"id": delivery["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    import secrets
    otp = f"{secrets.randbelow(1000000):06d}"
    expires = (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    await db.deliveries.update_one({"id": did}, {"$set": {
        "otp_hash": hash_password(otp),
        "otp_expires": expires,
        "otp_verified": False,
        "updated_at": now_iso(),
    }})
    await _log_whatsapp(lead, "delivery_otp", {"otp": otp, "expires_at": expires}, user)
    # Returned to sales exec UI to read aloud / share with customer
    return {"otp": otp, "expires_at": expires}


@api.post("/deliveries/{did}/otp-verify")
async def otp_verify(did: str, otp: str = Query(...), user: dict = Depends(get_current_user)):
    delivery = await db.deliveries.find_one({"id": did}, {"_id": 0})
    if not delivery:
        raise HTTPException(404, "Delivery not found")
    lead = await db.leads.find_one({"id": delivery["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")
    if not delivery.get("otp_hash") or not delivery.get("otp_expires"):
        raise HTTPException(400, "No OTP generated")
    if datetime.fromisoformat(delivery["otp_expires"]) < datetime.now(timezone.utc):
        raise HTTPException(400, "OTP expired")
    if not verify_password(otp, delivery["otp_hash"]):
        raise HTTPException(400, "Invalid OTP")
    await db.deliveries.update_one({"id": did}, {"$set": {
        "otp_verified": True,
        "otp_verified_at": now_iso(),
        "updated_at": now_iso(),
    }})
    return {"ok": True}


@api.post("/deliveries/{did}/complete")
async def complete_delivery(did: str, user: dict = Depends(get_current_user)):
    delivery = await db.deliveries.find_one({"id": did}, {"_id": 0})
    if not delivery:
        raise HTTPException(404, "Delivery not found")
    lead = await db.leads.find_one({"id": delivery["lead_id"]}, {"_id": 0})
    if not lead or not await can_access_lead(user, lead):
        raise HTTPException(403, "Access denied")

    # Payment completeness — allow if payment completed OR finance approved + downpayment received
    booking = await db.bookings.find_one({"id": delivery.get("booking_id")}, {"_id": 0}) if delivery.get("booking_id") else None
    if booking and float(booking.get("pending_amount") or 0) > 0.01:
        fc = await db.finance_cases.find_one({"lead_id": lead["id"]}, {"_id": 0})
        finance_ok = bool(fc and fc.get("status") == "Approved" and fc.get("downpayment_received"))
        if not finance_ok:
            raise HTTPException(400, f"Pending amount ₹{booking['pending_amount']} must be cleared (or finance approved + downpayment received) before delivery")

    checklist = delivery.get("checklist") or {}
    if not all([checklist.get("payment_completed"), checklist.get("documents_verified"),
                checklist.get("vehicle_ready"), checklist.get("accessories_ready")]):
        raise HTTPException(400, "All checklist items must be completed")
    if not delivery.get("otp_verified"):
        raise HTTPException(400, "Customer OTP verification required")

    # Documents: all mandatory-for-delivery types must be Verified
    docs = await db.documents.find({"lead_id": lead["id"], "is_latest": True}, {"_id": 0}).to_list(200)
    missing = _stage_doc_requirements("Delivered", docs)
    if missing:
        raise HTTPException(400, f"Missing verified documents: {', '.join(missing)}")

    await db.deliveries.update_one({"id": did}, {"$set": {
        "status": "Delivered",
        "delivered_at": now_iso(),
        "completed_by": user["id"],
        "updated_at": now_iso(),
    }})
    # Advance lead stage
    await db.leads.update_one({"id": lead["id"]},
                              {"$set": {"stage": "RTO", "updated_at": now_iso()}})
    await add_timeline(lead["id"], "Vehicle Delivered", user, {})
    await add_timeline(lead["id"], "Stage Changed", user,
                       {"from": "Delivered", "to": "RTO", "via": "delivery_complete"})
    # WhatsApp post-delivery intents
    await _log_whatsapp(lead, "delivery_thank_you",
                        {"message": f"Thanks for buying with us, {lead.get('customer_name')}!"}, user)
    await _log_whatsapp(lead, "feedback_reminder",
                        {"send_after_days": 5, "message": "How was your experience?"}, user)
    await _log_whatsapp(lead, "rc_followup",
                        {"send_after_days": 45, "message": "Your RC should be ready. Please collect."}, user)
    return await db.deliveries.find_one({"id": did}, {"_id": 0})


@api.get("/deliveries/{did}/challan", response_class=FastResponse)
async def delivery_challan(did: str, request: Request,
                           authorization: Optional[str] = Header(None),
                           auth: Optional[str] = Query(None)):
    # Accept query-param token so print window works
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization[7:]
    elif auth:
        token = auth
    else:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except Exception:
        raise HTTPException(401, "Invalid token")

    delivery = await db.deliveries.find_one({"id": did}, {"_id": 0})
    if not delivery:
        raise HTTPException(404, "Delivery not found")
    lead = await db.leads.find_one({"id": delivery["lead_id"]}, {"_id": 0})
    booking = await db.bookings.find_one({"id": delivery.get("booking_id")}, {"_id": 0}) if delivery.get("booking_id") else None
    branch = await db.branches.find_one({"id": delivery.get("branch_id")}, {"_id": 0})
    payments = await db.payments.find({"booking_id": booking["id"] if booking else ""}, {"_id": 0}).sort("date", 1).to_list(200) if booking else []

    accessories_rows = "".join(
        f"<tr><td>{a.get('name','')}</td><td style='text-align:center'>{a.get('quantity',1)}</td><td style='text-align:right'>₹{a.get('value',0)}</td></tr>"
        for a in (delivery.get("accessories") or [])
    )
    payment_rows = "".join(
        f"<tr><td>{p['date']}</td><td>{p['mode']}</td><td style='text-align:right'>₹{p['amount']}</td></tr>"
        for p in payments
    )

    html = f"""<!doctype html><html><head><meta charset="utf-8"><title>Delivery Challan</title>
<style>
  body {{ font-family: 'IBM Plex Sans', Arial, sans-serif; color: #09090b; padding: 40px; max-width: 800px; margin: auto; }}
  h1 {{ font-size: 24px; margin: 0 0 4px; letter-spacing: -0.02em; }}
  .overline {{ font-size: 10px; letter-spacing: 0.2em; color: #52525b; font-weight: 700; text-transform: uppercase; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 13px; }}
  th, td {{ border: 1px solid #e4e4e7; padding: 8px; text-align: left; }}
  th {{ background: #fafafa; font-size: 11px; letter-spacing: 0.1em; text-transform: uppercase; }}
  .two {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  .box {{ border: 1px solid #e4e4e7; padding: 16px; }}
  .muted {{ color: #52525b; font-size: 12px; }}
  .total {{ font-weight: 700; font-size: 16px; }}
  @media print {{ .noprint {{ display: none }} body {{ padding: 0 }} }}
</style></head><body>
  <div class="noprint" style="text-align:right;margin-bottom:16px">
    <button onclick="window.print()">Print</button>
  </div>
  <div style="border-bottom:2px solid #09090b; padding-bottom: 12px; margin-bottom: 16px;">
    <div class="overline">Torque Dealership CRM</div>
    <h1>Delivery Challan</h1>
    <div class="muted">Branch: {branch.get('name') if branch else '—'} · Date: {delivery.get('delivered_at','')[:10] or delivery.get('delivery_date','')}</div>
  </div>
  <div class="two">
    <div class="box">
      <div class="overline">Customer</div>
      <div style="margin-top:4px"><b>{delivery.get('customer_name','')}</b></div>
      <div class="muted">Phone: {delivery.get('customer_phone','')}</div>
      <div class="muted">Address: {lead.get('address') if lead else '—'}</div>
    </div>
    <div class="box">
      <div class="overline">Vehicle</div>
      <div style="margin-top:4px"><b>Chassis: {delivery.get('chassis_number','—')}</b></div>
      <div class="muted">Engine: {delivery.get('engine_number','—')}</div>
    </div>
  </div>
  <h3 style="margin-top:16px">Payment Summary</h3>
  <table>
    <tr><th>Date</th><th>Mode</th><th style="text-align:right">Amount</th></tr>
    {payment_rows or '<tr><td colspan="3" class="muted">No payments recorded</td></tr>'}
    <tr><td colspan="2" class="total">Total Paid</td><td class="total" style="text-align:right">₹{(booking or {}).get('total_paid',0)}</td></tr>
    <tr><td colspan="2" class="total">Final Deal Price</td><td class="total" style="text-align:right">₹{(booking or {}).get('final_deal_price',0)}</td></tr>
    <tr><td colspan="2" class="total">Pending</td><td class="total" style="text-align:right">₹{(booking or {}).get('pending_amount',0)}</td></tr>
  </table>
  <h3 style="margin-top:16px">Accessories</h3>
  <table>
    <tr><th>Item</th><th>Qty</th><th style="text-align:right">Value</th></tr>
    {accessories_rows or '<tr><td colspan="3" class="muted">No accessories</td></tr>'}
  </table>
  <div class="two" style="margin-top:48px">
    <div><div class="muted">Delivered by</div><div style="margin-top:40px;border-top:1px solid #09090b;padding-top:6px;font-weight:700">{delivery.get('created_by_name','')}</div></div>
    <div><div class="muted">Customer signature</div><div style="margin-top:40px;border-top:1px solid #09090b;padding-top:6px;font-weight:700">OTP Verified: {'✓' if delivery.get('otp_verified') else '—'}</div></div>
  </div>
</body></html>"""
    return FastResponse(content=html, media_type="text/html")




@api.get("/analytics/deals")
async def analytics_deals(user: dict = Depends(get_current_user)):
    base: Dict[str, Any] = {}
    if user["role"] == "sales_executive":
        base["assigned_to"] = user["id"]
    elif user["role"] == "admin":
        base["branch_id"] = user.get("branch_id")

    in_progress = await db.leads.count_documents({**base, "stage": "Booking"})
    booked = await db.leads.count_documents({**base, "stage": {"$in": ["Booking", "Delivered", "RTO", "Delivered"]}})
    total_with_deal = await db.leads.count_documents({**base, "deal.final_deal_price": {"$gt": 0}})
    deal_to_booking_rate = round((booked / (in_progress + booked)) * 100, 1) if (in_progress + booked) else 0.0

    # Loss reasons
    pipeline = [{"$match": {**base, "stage": "Lost"}},
                {"$group": {"_id": "$lost_reason", "count": {"$sum": 1}}}]
    loss_reasons = {}
    async for row in db.leads.aggregate(pipeline):
        loss_reasons[row["_id"] or "Unknown"] = row["count"]

    # Branch-wise deal value (super_admin only scope meaningful, but works for admin too)
    pipeline2 = [{"$match": {**base, "deal.final_deal_price": {"$gt": 0}}},
                 {"$group": {"_id": "$branch_id",
                             "count": {"$sum": 1},
                             "total_value": {"$sum": "$deal.final_deal_price"}}}]
    branches = []
    async for row in db.leads.aggregate(pipeline2):
        branches.append({"branch_id": row["_id"], "count": row["count"], "total_value": row["total_value"]})

    return {
        "in_progress": in_progress,
        "booked": booked,
        "total_with_deal": total_with_deal,
        "deal_to_booking_rate": deal_to_booking_rate,
        "loss_reasons": loss_reasons,
        "branches": branches,
    }


# ============================================================
# Admin — Production data purge
# ============================================================

@api.get("/admin/export-data")
async def export_data(user: dict = Depends(get_current_user)):
    """Export ALL collections as a single JSON file. Super-admin only.
    Returns a downloadable JSON with all data, suitable for backup or migration.
    Excludes MongoDB `_id` and `password_hash` for safety.
    """
    if user.get("role") != "super_admin":
        raise HTTPException(403, "Only super admin can export data")

    import json as _json
    collections = [
        "users", "branches", "leads",
        "followups", "bookings", "deliveries", "allotments", "payments",
        "documents", "files", "exchange_valuations", "negotiation_history",
        "timeline", "wa_messages", "wa_templates",
        "campaigns", "automation_rules", "finance_cases", "inventory",
        "brands", "vehicle_models", "variants", "colors",
        "audit_logs", "reminders", "system_flags",
    ]
    export: Dict[str, Any] = {
        "exported_at": now_iso(),
        "exported_by": {"id": user["id"], "name": user["name"], "email": user.get("email")},
        "db_name": DB_NAME,
        "collections": {},
        "counts": {},
    }
    for coll in collections:
        try:
            docs = await db[coll].find({}, {"_id": 0, "password_hash": 0}).to_list(None)
            # datetime → iso strings for clean JSON
            export["collections"][coll] = docs
            export["counts"][coll] = len(docs)
        except Exception as e:
            export["counts"][coll] = f"ERR: {e}"

    payload = _json.dumps(export, ensure_ascii=False, indent=2, default=str)
    fname = f"servall_crm_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    await log_audit(user, "data_exported", entity_type="system",
                    meta={"counts": export["counts"]})
    return FastResponse(
        content=payload,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )

@api.post("/admin/purge-demo-data")
async def purge_demo_data(
    confirm: str = "",
    keep_master_data: bool = True,
    keep_users: bool = False,
    user: dict = Depends(get_current_user),
):
    """DESTRUCTIVE. Super-admin only. Wipes all transactional data for production go-live.

    Flags:
      keep_master_data=True (default)  → preserves brands/models/variants/colors
      keep_users=True                   → preserves all users (admin + sales execs); only wipes
                                          transactional collections (leads, bookings, etc.)
    """
    if user.get("role") != "super_admin":
        raise HTTPException(403, "Only super admin can purge data")
    if confirm != "SERVALL_PURGE":
        raise HTTPException(400, "Missing confirm token — pass confirm=SERVALL_PURGE")

    stats: Dict[str, int] = {}
    transactional = [
        "leads", "followups", "bookings", "deliveries", "allotments", "payments",
        "files", "documents", "exchange_valuations", "negotiation_history", "timeline",
        "campaigns", "automation_rules", "wa_messages", "whatsapp_logs",
        "audit_logs", "finance_cases", "inventory",
    ]
    for col in transactional:
        res = await db[col].delete_many({})
        stats[col] = res.deleted_count

    # Delete non-super-admin users UNLESS keep_users is set
    if not keep_users:
        res = await db.users.delete_many({"role": {"$ne": "super_admin"}})
        stats["users_deleted"] = res.deleted_count
    else:
        stats["users_kept"] = await db.users.count_documents({})

    # Optionally wipe master data
    if not keep_master_data:
        for col in ("brands", "vehicle_models", "variants", "colors"):
            res = await db[col].delete_many({})
            stats[f"{col}_deleted"] = res.deleted_count

    # Set production_mode flag so seed doesn't recreate sample data on restart
    await db.system_flags.update_one(
        {"id": "production_mode"},
        {"$set": {
            "id": "production_mode",
            "enabled": True,
            "purged_by": user["id"],
            "purged_by_name": user["name"],
            "purged_at": now_iso(),
            "kept_users": keep_users,
        }},
        upsert=True,
    )

    return {
        "ok": True,
        "message": "Demo data purged. System is now in production mode.",
        "stats": stats,
    }




# ============================================================
# Seed & startup

# ============================================================
# Inventory / Stock — chassis-level vehicle stock
# ============================================================

class InventoryItem(BaseModel):
    brand: str
    model: str
    variant: Optional[str] = None
    color: Optional[str] = None
    chassis_number: str
    engine_number: Optional[str] = None
    branch_id: Optional[str] = None  # null = central / Bilimora hub
    notes: Optional[str] = None


@api.get("/inventory")
async def list_inventory(
    status: Optional[str] = None,
    brand: Optional[str] = None,
    model: Optional[str] = None,
    variant: Optional[str] = None,
    color: Optional[str] = None,
    chassis: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    q: Dict[str, Any] = {}
    if status:
        q["status"] = status
    if brand:
        q["brand"] = {"$regex": f"^{brand}$", "$options": "i"}
    if model:
        q["model"] = {"$regex": f"^{model}$", "$options": "i"}
    if variant:
        q["variant"] = {"$regex": f"^{variant}$", "$options": "i"}
    if color:
        q["color"] = {"$regex": f"^{color}$", "$options": "i"}
    if chassis:
        q["chassis_number"] = {"$regex": chassis, "$options": "i"}
    items = []
    async for it in db.inventory.find(q, {"_id": 0}).sort("created_at", -1).limit(2000):
        items.append(it)
    return items


@api.post("/inventory")
async def add_inventory(body: InventoryItem, user: dict = Depends(get_current_user)):
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(403, "Only admin/super-admin can add stock")
    chassis = (body.chassis_number or "").strip().upper()
    if not chassis:
        raise HTTPException(400, "Chassis number required")
    dup = await db.inventory.find_one({"chassis_number": chassis}, {"_id": 0})
    if dup:
        raise HTTPException(400, f"Chassis {chassis} already exists in stock")
    rec = {
        "id": str(uuid.uuid4()),
        "brand": body.brand.strip(),
        "model": body.model.strip(),
        "variant": (body.variant or "").strip() or None,
        "color": (body.color or "").strip() or None,
        "chassis_number": chassis,
        "engine_number": (body.engine_number or "").strip() or None,
        "branch_id": body.branch_id,
        "status": "available",
        "booked_lead_id": None,
        "booked_booking_id": None,
        "notes": body.notes,
        "created_by": user["id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.inventory.insert_one(rec)
    rec.pop("_id", None)
    return rec


@api.delete("/inventory/{iid}")
async def delete_inventory(iid: str, user: dict = Depends(get_current_user)):
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(403, "Only admin/super-admin can delete stock")
    item = await db.inventory.find_one({"id": iid}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Inventory item not found")
    if item.get("status") == "booked":
        raise HTTPException(400, "Cannot delete a booked vehicle — cancel its booking first")
    await db.inventory.delete_one({"id": iid})
    return {"ok": True}


@api.post("/inventory/upload")
async def upload_inventory_excel(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Bulk-upload inventory from CSV/Excel.
    Required columns (case-insensitive): brand, model, chassis_number
    Optional: variant, color, engine_number, branch_id, notes
    Returns: {ok, added, skipped_duplicates, errors}
    """
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(403, "Only admin/super-admin can upload stock")
    raw = await file.read()
    fname = (file.filename or "").lower()
    rows: List[Dict[str, Any]] = []
    try:
        if fname.endswith(".csv") or "csv" in (file.content_type or ""):
            import csv as _csv
            import io as _io
            txt = raw.decode("utf-8", errors="ignore")
            reader = _csv.DictReader(_io.StringIO(txt))
            rows = [{(k or "").strip().lower(): (v or "").strip() for k, v in r.items()} for r in reader]
        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            try:
                from openpyxl import load_workbook
                import io as _io
                wb = load_workbook(filename=_io.BytesIO(raw), read_only=True)
                ws = wb.active
                headers = []
                for ridx, row in enumerate(ws.iter_rows(values_only=True)):
                    if ridx == 0:
                        headers = [str(c or "").strip().lower() for c in row]
                    else:
                        d = {headers[i]: (str(c).strip() if c is not None else "") for i, c in enumerate(row) if i < len(headers)}
                        if any(d.values()):
                            rows.append(d)
            except ImportError:
                raise HTTPException(400, "Excel support unavailable — install openpyxl, or upload CSV instead")
        else:
            raise HTTPException(400, "Upload .csv or .xlsx file")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    added = 0
    skipped = 0
    errors: List[str] = []
    for i, r in enumerate(rows, start=2):  # row 2 = first data row in spreadsheet
        chassis = (r.get("chassis_number") or r.get("chassis") or "").strip().upper()
        brand = (r.get("brand") or "").strip()
        model = (r.get("model") or "").strip()
        if not chassis or not brand or not model:
            errors.append(f"Row {i}: missing brand/model/chassis_number")
            continue
        dup = await db.inventory.find_one({"chassis_number": chassis}, {"_id": 0})
        if dup:
            skipped += 1
            continue
        rec = {
            "id": str(uuid.uuid4()),
            "brand": brand, "model": model,
            "variant": r.get("variant") or None,
            "color": r.get("color") or None,
            "chassis_number": chassis,
            "engine_number": (r.get("engine_number") or r.get("engine") or "").strip() or None,
            "branch_id": r.get("branch_id") or None,
            "status": "available",
            "booked_lead_id": None,
            "booked_booking_id": None,
            "notes": r.get("notes") or None,
            "created_by": user["id"],
            "created_at": now_iso(),
            "updated_at": now_iso(),
        }
        await db.inventory.insert_one(rec)
        added += 1
    return {"ok": True, "added": added, "skipped_duplicates": skipped, "errors": errors[:50]}


# Migrate legacy stage names → new simplified stages (idempotent)
@api.post("/admin/migrate-stages")
async def migrate_stages(user: dict = Depends(get_current_user)):
    if user.get("role") != "super_admin":
        raise HTTPException(403, "Only super admin")
    counts: Dict[str, int] = {}
    for old, new in STAGE_ALIAS.items():
        res = await db.leads.update_many({"stage": old}, {"$set": {"stage": new, "updated_at": now_iso()}})
        counts[f"{old}->{new}"] = res.modified_count
    return {"ok": True, "migrated": counts}



# ============================================================

async def seed_data():
    # If production_mode flag is set (via purge-demo-data), skip ALL demo seeding
    prod_flag = await db.system_flags.find_one({"id": "production_mode"}, {"_id": 0})
    is_production = bool(prod_flag and prod_flag.get("enabled"))

    # Indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.leads.create_index("id", unique=True)
    await db.leads.create_index("assigned_to")
    await db.leads.create_index("branch_id")
    await db.timeline.create_index("lead_id")
    await db.followups.create_index("lead_id")
    await db.bookings.create_index("lead_id")
    await db.bookings.create_index("branch_id")
    await db.payments.create_index("booking_id")
    await db.allotments.create_index("chassis_number", unique=True)
    await db.allotments.create_index("booking_id", unique=True)
    await db.documents.create_index([("lead_id", 1), ("doc_type", 1)])
    await db.documents.create_index([("doc_type", 1), ("doc_number", 1)])
    await db.deliveries.create_index("lead_id")
    await db.whatsapp_logs.create_index("lead_id")
    await db.finance_cases.create_index("lead_id", unique=True)
    await db.exchange_valuations.create_index("lead_id")
    await db.wa_templates.create_index("name", unique=True)
    await db.automation_rules.create_index("event")
    await db.wa_messages.create_index("lead_id")
    await db.wa_messages.create_index("campaign_id")
    await db.wa_optouts.create_index("lead_id", unique=True)
    await db.campaigns.create_index("status")
    await db.audit_logs.create_index([("created_at", -1)])
    await db.audit_logs.create_index("actor_id")
    await db.audit_logs.create_index("branch_id")
    await db.audit_logs.create_index("action")
    # Sparse unique so legacy records without phone / code still work
    try:
        await db.users.create_index("phone", unique=True, sparse=True)
    except Exception as e:
        logger.warning(f"users.phone index failed: {e}")
    try:
        await db.branches.create_index("code", unique=True, sparse=True)
    except Exception as e:
        logger.warning(f"branches.code index failed: {e}")

    # Branches
    branch_defs = [
        {"name": "Bilimora", "code": "BLM", "city": "Bilimora"},
        {"name": "Chikhli", "code": "CHK", "city": "Chikhli"},
        {"name": "Amalsad", "code": "AMD", "city": "Amalsad"},
        {"name": "Vansda", "code": "VNS", "city": "Vansda"},
        {"name": "Gandevi", "code": "GND", "city": "Gandevi"},
    ]
    branches = {}
    for b in branch_defs:
        existing = await db.branches.find_one({"name": b["name"]}, {"_id": 0})
        if existing:
            # Backfill new fields on older records
            patch = {}
            if existing.get("is_active") is None:
                patch["is_active"] = True
            if existing.get("allow_login_when_inactive") is None:
                patch["allow_login_when_inactive"] = True
            if not existing.get("code"):
                patch["code"] = b["code"]
            if not existing.get("city"):
                patch["city"] = b["city"]
            if patch:
                await db.branches.update_one({"id": existing["id"]}, {"$set": patch})
                existing.update(patch)
            branches[b["name"]] = existing
        else:
            doc = {
                "id": str(uuid.uuid4()), "name": b["name"], "code": b["code"],
                "city": b["city"], "address": None,
                "assigned_admin_id": None,
                "is_active": True,
                "allow_login_when_inactive": True,
                "created_at": now_iso(),
            }
            await db.branches.insert_one(doc)
            branches[b["name"]] = doc

    # Brands & Models & Variants
    brand_defs = {
        "Honda": {
            "Activa 6G": ["Standard", "DLX", "Smart"],
            "Shine": ["Drum", "Disc"],
            "Dio": ["Standard", "DLX"],
        },
        "Hero": {
            "Splendor Plus": ["Kick", "Self", "iSmart"],
            "HF Deluxe": ["Kick", "Self"],
        },
        "TVS": {
            "Jupiter": ["Standard", "ZX", "Classic"],
            "Apache RTR 160": ["Drum", "Disc"],
        },
        "Suzuki": {
            "Access 125": ["Standard", "Special Edition"],
            "Burgman Street": ["125", "EX"],
        },
    }
    for bname, models in brand_defs.items():
        bdoc = await db.brands.find_one({"name": bname}, {"_id": 0})
        if not bdoc:
            bdoc = {"id": str(uuid.uuid4()), "name": bname, "created_at": now_iso()}
            await db.brands.insert_one(bdoc)
        for mname, variants in models.items():
            mdoc = await db.vehicle_models.find_one({"name": mname, "brand_id": bdoc["id"]}, {"_id": 0})
            if not mdoc:
                mdoc = {"id": str(uuid.uuid4()), "name": mname, "brand_id": bdoc["id"], "created_at": now_iso()}
                await db.vehicle_models.insert_one(mdoc)
            for vname in variants:
                if not await db.variants.find_one({"name": vname, "model_id": mdoc["id"]}):
                    await db.variants.insert_one({
                        "id": str(uuid.uuid4()), "name": vname,
                        "model_id": mdoc["id"], "created_at": now_iso()
                    })

    # Colors
    color_defs = [
        ("Pearl White", "#F5F5F5"),
        ("Matte Black", "#222222"),
        ("Racing Red", "#E11D48"),
        ("Ocean Blue", "#1E40AF"),
        ("Silver", "#C0C0C0"),
    ]
    for cname, hex_code in color_defs:
        if not await db.colors.find_one({"name": cname}):
            await db.colors.insert_one({
                "id": str(uuid.uuid4()), "name": cname, "hex": hex_code,
                "created_at": now_iso()
            })

    # Users
    user_defs = [
        {"email": "superadmin@dealer.com", "password": "super123", "name": "Super Admin",
         "phone": "9000000000", "role": "super_admin", "branch": None, "manager": None},
    ]
    if not is_production:
        user_defs += [
        {"email": "admin@dealer.com", "password": "admin123", "name": "Ravi Admin",
         "phone": "9000000001", "role": "admin", "branch": "Bilimora",
         "manager": "superadmin@dealer.com"},
        {"email": "sales1@dealer.com", "password": "sales123", "name": "Priya Sales",
         "phone": "9000000011", "role": "sales_executive", "branch": "Bilimora",
         "manager": "admin@dealer.com"},
        {"email": "sales2@dealer.com", "password": "sales123", "name": "Amit Sales",
         "phone": "9000000012", "role": "sales_executive", "branch": "Bilimora",
         "manager": "admin@dealer.com"},
        {"email": "sales3@dealer.com", "password": "sales123", "name": "Neha Sales",
         "phone": "9000000013", "role": "sales_executive", "branch": "Chikhli",
         "manager": "admin@dealer.com"},
        {"email": "sales4@dealer.com", "password": "sales123", "name": "Vikram Sales",
         "phone": "9000000014", "role": "sales_executive", "branch": "Gandevi",
         "manager": "admin@dealer.com"},
        ]
    # First pass: upsert users (without manager resolved)
    for u in user_defs:
        existing = await db.users.find_one({"email": u["email"]})
        branch_id = branches[u["branch"]]["id"] if u["branch"] else None
        base = {
            "name": u["name"],
            "phone": u["phone"],
            "role": u["role"],
            "branch_id": branch_id,
            "is_active": True,
            "joining_date": existing.get("joining_date") if existing else "2025-01-01",
            "permissions": (existing or {}).get("permissions") or {},
        }
        if not existing:
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "email": u["email"],
                "password_hash": hash_password(u["password"]),
                "created_at": now_iso(),
                "joining_date": "2025-01-01",
                **{k: v for k, v in base.items() if k != "joining_date"},
            })
        else:
            # Ensure password matches seed and backfill new fields
            patch = dict(base)
            if not verify_password(u["password"], existing["password_hash"]):
                patch["password_hash"] = hash_password(u["password"])
            await db.users.update_one({"email": u["email"]}, {"$set": patch})

    # Second pass: resolve reporting_manager_id by email
    email_to_id: Dict[str, str] = {}
    for u in user_defs:
        doc = await db.users.find_one({"email": u["email"]}, {"_id": 0})
        if doc:
            email_to_id[u["email"]] = doc["id"]
    for u in user_defs:
        mgr_id = email_to_id.get(u.get("manager")) if u.get("manager") else None
        await db.users.update_one({"email": u["email"]},
                                  {"$set": {"reporting_manager_id": mgr_id}})

    # Assign Bilimora admin to branch
    bilimora = branches.get("Bilimora")
    admin_id = email_to_id.get("admin@dealer.com")
    if bilimora and admin_id:
        await db.branches.update_one({"id": bilimora["id"]},
                                     {"$set": {"assigned_admin_id": admin_id}})

    # Seed default WA templates
    default_templates = [
        ("inquiry_welcome", "inquiry",
         "Hi {{customer_name}}! Thanks for your interest in {{vehicle}} at {{branch}}. I'm {{sales_exec}} — reply with any questions, or drop by the showroom anytime."),
        ("followup_reminder", "followup",
         "Hi {{customer_name}}, checking in on your interest in {{vehicle}}. Would you like to schedule a test ride?"),
        ("deal_offer", "deal",
         "Hi {{customer_name}}! Here's your offer for {{vehicle}} — final price Rs.{{deal_amount}}, discount Rs.{{discount}}. Let me know if you'd like to proceed."),
        ("booking_confirm", "booking",
         "Your booking for {{vehicle}} is confirmed at {{branch}}! Our team will keep you posted on delivery."),
        ("delivery_thankyou", "delivery",
         "Congratulations {{customer_name}}! Your {{vehicle}} is delivered. Enjoy the ride and reach out anytime."),
        ("feedback_5d", "feedback",
         "Hi {{customer_name}}! How has your {{vehicle}} been so far? We'd love your feedback."),
        ("rc_reminder_45d", "reminder",
         "Hi {{customer_name}}, your RC for the {{vehicle}} should be ready. Please visit {{branch}} to collect."),
        ("lost_reengage", "reengage",
         "Hi {{customer_name}}, still thinking about {{vehicle}}? We have a new offer that might change your mind. Reply YES to know more."),
    ]
    for name, cat, body in default_templates:
        existing = await db.wa_templates.find_one({"name": name})
        if not existing:
            await db.wa_templates.insert_one({
                "id": str(uuid.uuid4()), "name": name, "category": cat,
                "message_type": "text", "body": body, "media_url": None,
                "active": True, "created_at": now_iso(), "updated_at": now_iso(),
            })

    # Sample leads
    count = await db.leads.count_documents({})
    if count == 0 and not is_production:
        sales_exec = await db.users.find_one({"email": "sales1@dealer.com"}, {"_id": 0})
        honda = await db.brands.find_one({"name": "Honda"}, {"_id": 0})
        activa = await db.vehicle_models.find_one({"name": "Activa 6G"}, {"_id": 0})
        sample_leads = [
            {"customer_name": "Rakesh Patel", "phone": "9876543210", "source": "Walk-in",
             "priority": "Hot", "stage": "Follow-up"},
            {"customer_name": "Meera Shah", "phone": "9876501234", "source": "WhatsApp",
             "priority": "Warm", "stage": "Follow-up"},
            {"customer_name": "Arjun Desai", "phone": "9845012345", "source": "Digital Marketing",
             "priority": "Cold", "stage": "Inquiry"},
            {"customer_name": "Kavita Joshi", "phone": "9812345678", "source": "Referral",
             "priority": "Hot", "stage": "Test Ride"},
            {"customer_name": "Sanjay Modi", "phone": "9823456789", "source": "Tele-in",
             "priority": "Warm", "stage": "Booking",
             "deal_c": 78000, "deal_o": 75000},
        ]
        for s in sample_leads:
            lid = str(uuid.uuid4())
            await db.leads.insert_one({
                "id": lid,
                "customer_name": s["customer_name"],
                "phone": s["phone"],
                "alt_phone": None, "birthdate": None, "address": "Gujarat",
                "source": s["source"],
                "branch_id": branches["Bilimora"]["id"],
                "priority": s["priority"],
                "assigned_to": sales_exec["id"] if sales_exec else None,
                "brand_id": honda["id"] if honda else None,
                "model_id": activa["id"] if activa else None,
                "variant_id": None, "color_id": None,
                "purchase_type": "New Purchase",
                "exchange": None,
                "deal": {"customer_expected_price": s.get("deal_c"),
                         "offered_price": s.get("deal_o"),
                         "discount": None, "interest_level": s["priority"]},
                "payment_mode": None, "finance": None,
                "next_followup_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "next_followup_type": "Call",
                "notes": "Sample lead",
                "stage": s["stage"],
                "created_by": sales_exec["id"] if sales_exec else None,
                "created_at": now_iso(),
                "updated_at": now_iso(),
                "followup_count": 0,
                "documents": [],
                "lost_reason": None, "lost_reason_text": None,
            })
            await db.timeline.insert_one({
                "id": str(uuid.uuid4()), "lead_id": lid,
                "event": "Lead Created", "meta": {"source": s["source"]},
                "actor_id": sales_exec["id"] if sales_exec else None,
                "actor_name": sales_exec["name"] if sales_exec else "System",
                "created_at": now_iso(),
            })

    # Seed default WhatsApp templates + automation rules (idempotent)
    default_templates = [
        {
            "key_name": "Inquiry — send catalog",
            "body": "Hi {{customer_name}}, thanks for your interest at Servall! View our catalog: https://servall.com/catalog",
            "category": "inquiry",
        },
        {
            "key_name": "Delivery — thank you",
            "body": "Congratulations {{customer_name}}! Your vehicle has been delivered. Thank you for choosing Servall — safe & happy rides! 🏍️",
            "category": "delivery",
        },
        {
            "key_name": "Feedback — request",
            "body": "Hi {{customer_name}}, we'd love your feedback on your recent purchase. Reply here or visit: https://servall.com/feedback",
            "category": "feedback",
        },
    ]
    tpl_by_name = {}
    for td in default_templates:
        existing = await db.wa_templates.find_one({"name": td["key_name"]}, {"_id": 0})
        if existing:
            tpl_by_name[td["key_name"]] = existing["id"]
        else:
            tid = str(uuid.uuid4())
            await db.wa_templates.insert_one({
                "id": tid,
                "name": td["key_name"],
                "category": td["category"],
                "message_type": "text",
                "body": td["body"],
                "active": True,
                "created_at": now_iso(),
            })
            tpl_by_name[td["key_name"]] = tid

    default_rules = [
        {"name": "Auto-send catalog on inquiry", "event": "inquiry_created", "template_name": "Inquiry — send catalog", "delay": 0},
        {"name": "Thank you on delivery", "event": "delivery_completed", "template_name": "Delivery — thank you", "delay": 0},
        {"name": "Feedback reminder", "event": "feedback_reminder", "template_name": "Feedback — request", "delay": 0},
    ]
    for rd in default_rules:
        existing = await db.automation_rules.find_one({"name": rd["name"]}, {"_id": 0})
        if existing:
            continue
        tid = tpl_by_name.get(rd["template_name"])
        if not tid:
            continue
        await db.automation_rules.insert_one({
            "id": str(uuid.uuid4()),
            "name": rd["name"],
            "event": rd["event"],
            "conditions": {},
            "template_id": tid,
            "delay_minutes": rd["delay"],
            "active": True,
            "created_at": now_iso(),
        })


@app.on_event("startup")
async def on_startup():
    try:
        await seed_data()
        logger.info("Seed complete")
    except Exception as e:
        logger.exception(f"Seed failed: {e}")
    try:
        init_storage()
        logger.info("Object storage initialized")
    except Exception as e:
        logger.error(f"Storage init error: {e}")


@app.on_event("shutdown")
async def on_shutdown():
    client.close()


# ============================================================
# Constants endpoint
# ============================================================

@api.get("/constants")
async def get_constants():
    return {
        "lead_sources": LEAD_SOURCES,
        "priorities": PRIORITIES,
        "stages": STAGES,
        "payment_modes": PAYMENT_MODES,
        "followup_types": FOLLOWUP_TYPES,
        "call_statuses": CALL_STATUSES,
        "customer_responses": CUSTOMER_RESPONSES,
        "outcome_tags": OUTCOME_TAGS,
        "lost_reasons": LOST_REASONS,
        "customer_types": CUSTOMER_TYPES,
        "deal_loss_reasons": DEAL_LOSS_REASONS,
        "deal_statuses": DEAL_STATUSES,
        "booking_statuses": BOOKING_STATUSES,
        "allotment_statuses": ALLOTMENT_STATUSES,
        "loan_statuses": LOAN_STATUSES,
        "doc_types": DOC_TYPES,
        "doc_statuses": DOC_STATUSES,
        "doc_requirements": DOC_REQUIREMENTS,
        "delivery_statuses": DELIVERY_STATUSES,
        "default_accessories": DEFAULT_ACCESSORIES,
        "payment_types": PAYMENT_TYPES,
        "payment_statuses": PAYMENT_STATUSES,
        "finance_statuses": FINANCE_STATUSES,
        "exchange_conditions": EXCHANGE_CONDITIONS,
        "wa_events": WA_EVENTS,
        "wa_message_types": WA_MESSAGE_TYPES,
        "wa_statuses": WA_STATUSES,
        "wa_reply_tags": WA_REPLY_TAGS,
        "campaign_types": CAMPAIGN_TYPES,
        "campaign_statuses": CAMPAIGN_STATUSES,
        "roles": ROLES,
        "config": {
            "discount_approval_threshold": DISCOUNT_APPROVAL_THRESHOLD,
            "followup_min_gap_seconds": FOLLOWUP_MIN_GAP_SECONDS,
            "sla_hours_no_followup": SLA_HOURS_NO_FOLLOWUP,
            "at_risk_missed_count": AT_RISK_MISSED_COUNT,
        },
    }


@api.get("/")
async def root():
    return {"service": "twowheeler-crm", "status": "ok"}


app.include_router(api)

# CORS — allow listed origins + a permissive regex for production custom domains.
# FRONTEND_URL env can be a comma-separated list for multiple explicit origins.
_frontend_urls = [u.strip() for u in (FRONTEND_URL or "").split(",") if u.strip()]
_cors_allow = list(set(_frontend_urls + [
    "http://localhost:3000",
    "http://localhost:5173",
]))

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_allow,
    allow_origin_regex=r"https?://(localhost(:\d+)?|.*\.servall\.in|.*\.emergent\.host|.*\.emergentagent\.com|.*\.preview\.emergentagent\.com)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
